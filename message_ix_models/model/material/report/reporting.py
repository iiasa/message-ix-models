# -*- coding: utf-8 -*-
"""
Created on Mon Mar  8 12:58:21 2021
This code produces the follwoing outputs:
message_ix_reporting.xlsx: message_ix level reporting
check.xlsx: can be used for checking the filtered variables
New_Reporting_Model_Scenario.xlsx: Reporting including the material variables
Merged_Model_Scenario.xlsx: Includes all IAMC variables
Material_global_grpahs.pdf

@author: unlu
"""

# NOTE: Works with pyam-iamc version 0.9.0
# Problems with the most recent versions:
# 0.11.0 --> Filtering with asterisk * does not work, dataframes are empty.
# Problems with emission and region filtering.
# https://github.com/IAMconsortium/pyam/issues/517
# 0.10.0 --> Plotting gives the follwoing error:
# ValueError: Can not plot data that does not extend for the entire year range.
# There are various tech.s that only have data starting from 2025 or 2030.
# foil_imp AFR, frunace_h2_aluminum, h2_i...

# PACKAGES

from ixmp import Platform
from message_ix import Scenario
from message_ix.reporting import Reporter
from ixmp.reporting import configure
from message_ix_models import ScenarioInfo
# from message_data.tools.post_processing.iamc_report_hackathon import report as reporting
# from message_data.model.material.util import read_config

import pandas as pd
import numpy as np
import pyam
import xlsxwriter
import os
import openpyxl

import plotly.graph_objects as go
import matplotlib

matplotlib.use("Agg")
from matplotlib import pyplot as plt

from pyam.plotting import OUTSIDE_LEGEND
from matplotlib.backends.backend_pdf import PdfPages
from message_ix_models.util import (
    package_data_path,
)

def print_full(x):
    pd.set_option("display.max_rows", len(x))
    print(x)
    pd.reset_option("display.max_rows")


def change_names(s):

    """Change the sector names according to IMAC format."""

    if s == "aluminum":
        s = "Non-Ferrous Metals|Aluminium"
    elif s == "steel":
        s = "Steel"
    elif s == "cement":
        s = "Non-Metallic Minerals|Cement"
    elif s == "petro":
        s = "Chemicals|High Value Chemicals"
    elif s == 'ammonia':
        s = "Chemicals|Ammonia"
    elif s == 'methanol':
        s = "Chemicals|Methanol"
    elif s == "BCA":
        s = "BC"
    elif s == "OCA":
        s = "OC"
    elif s == "CO2_industry":
        s == "CO2"
    else:
        s == s
    return s


def fix_excel(path_temp, path_new):

    """
    Fix the names of the regions or variables to be compatible
    with IAMC format. This is done in the final reported excel file
    (path_temp) and written to a new excel file (path_new).
    """
    # read Excel file and sheet by name
    workbook = openpyxl.load_workbook(path_temp)
    sheet = workbook["data"]

    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook['Sheet']
    new_sheet.title = 'data'
    new_sheet = new_workbook.active

    replacement = {
        "CO2_industry": "CO2",
        "R11_AFR|R11_AFR": "R11_AFR",
        "R11_CPA|R11_CPA": "R11_CPA",
        "R11_MEA|R11_MEA": "R11_MEA",
        "R11_FSU|R11_FSU": "R11_FSU",
        "R11_PAS|R11_PAS": "R11_PAS",
        "R11_SAS|R11_SAS": "R11_SAS",
        "R11_LAM|R11_LAM": "R11_LAM",
        "R11_NAM|R11_NAM": "R11_NAM",
        "R11_PAO|R11_PAO": "R11_PAO",
        "R11_EEU|R11_EEU": "R11_EEU",
        "R11_WEU|R11_WEU": "R11_WEU",
        "R11_AFR": "R11_AFR",
        "R11_CPA": "R11_CPA",
        "R11_MEA": "R11_MEA",
        "R11_FSU": "R11_FSU",
        "R11_PAS": "R11_PAS",
        "R11_SAS": "R11_SAS",
        "R11_LAM": "R11_LAM",
        "R11_NAM": "R11_NAM",
        "R11_PAO": "R11_PAO",
        "R11_EEU": "R11_EEU",
        "R11_WEU": "R11_WEU",
        "R12_AFR|R12_AFR": "R12_AFR",
        "R12_RCPA|R12_RCPA": "R12_RCPA",
        "R12_MEA|R12_MEA": "R12_MEA",
        "R12_FSU|R12_FSU": "R12_FSU",
        "R12_PAS|R12_PAS": "R12_PAS",
        "R12_SAS|R12_SAS": "R12_SAS",
        "R12_LAM|R12_LAM": "R12_LAM",
        "R12_NAM|R12_NAM": "R12_NAM",
        "R12_PAO|R12_PAO": "R12_PAO",
        "R12_EEU|R12_EEU": "R12_EEU",
        "R12_WEU|R12_WEU": "R12_WEU",
        "R12_CHN|R12_CHN": "R12_CHN",
        "R12_AFR": "R12_AFR",
        "R12_RCPA": "R12_RCPA",
        "R12_MEA": "R12_MEA",
        "R12_FSU": "R12_FSU",
        "R12_PAS": "R12_PAS",
        "R12_SAS": "R12_SAS",
        "R12_LAM": "R12_LAM",
        "R12_NAM": "R12_NAM",
        "R12_PAO": "R12_PAO",
        "R12_EEU": "R12_EEU",
        "R12_WEU": "R12_WEU",
        "R12_CHN": "R12_CHN",
        "World": "R12_GLB",
        "model": "Model",
        "scenario": "Scenario",
        "variable": "Variable",
        "region": "Region",
        "unit": "Unit",
        "BCA": "BC",
        "OCA": "OC",
    }
    # Iterate over the rows and replace
    for i in range(1, ((sheet.max_row) + 1)):
        data = [sheet.cell(row=i, column=col).value for col in range(1, ((sheet.max_column) + 1))]
        for index, value in enumerate(data):
            col_no = index + 1
            if value in replacement.keys():
                new_sheet.cell(row=i, column=col_no).value = replacement.get(value)
            else:
                new_sheet.cell(row=i, column=col_no).value = value

    new_workbook.save(path_new)

def report(context,scenario):

    # Obtain scenario information and directory

    s_info = ScenarioInfo(scenario)

    # In order to avoid confusion in the second reporting stage there should
    # no existing timeseries uploaded in the scenairo. Clear these except the
    # residential and commercial ones since they should be always included.

    # Activate this part to keep the residential and commercial variables
    # when the model is run with the buildigns linkage.
    # df_rem = df_rem[~(df_rem["variable"].str.contains("Residential") | \
    # df_rem["variable"].str.contains("Commercial"))]

    years = s_info.Y
    nodes = []
    for n in s_info.N:
        n = n + "*"
        nodes.append(n)

    if "R11_GLB*" in nodes:
        nodes.remove("R11_GLB*")
    elif "R12_GLB*" in nodes:
        nodes.remove("R12_GLB*")

    # Path for materials reporting output
    directory = package_data_path("material", "reporting_output")
    directory.mkdir(exist_ok=True)

    # Generate message_ix level reporting and dump to an excel file.

    rep = Reporter.from_scenario(scenario)
    configure(units={"replace": {"-": ""}})
    df = rep.get("message::default")
    name = os.path.join(directory, f"message_ix_reporting_{scenario.scenario}.xlsx")
    df.to_excel(name)
    print("message_ix level reporting generated")

    # Obtain a pyam dataframe / filter / global aggregation

    path = os.path.join(directory, f"message_ix_reporting_{scenario.scenario}.xlsx")
    report = pd.read_excel(path)
    report.Unit.fillna("", inplace=True)
    df = pyam.IamDataFrame(report)
    df.filter(region=nodes, year=years, inplace=True)
    df.filter(
        variable=[
            "out|new_scrap|aluminum|*",
            "out|final_material|aluminum|prebake_aluminum|M1",
            "out|final_material|aluminum|secondary_aluminum|M1",
            "out|final_material|aluminum|soderberg_aluminum|M1",
            "out|new_scrap|steel|*",
            "in|new_scrap|steel|*",
            "out|final_material|steel|*",
            "out|useful_material|steel|import_steel|*",
            "out|secondary_material|NH3|*",
            "out|primary|methanol|*",
            "out|primary_material|methanol|*",
            "out|final|methanol|*",
            "out|final_material|methanol|*",
            "in|useful_material|steel|export_steel|*",
            "out|useful_material|aluminum|import_aluminum|*",
            "in|useful_material|aluminum|export_aluminum|*",
            "out|final_material|*|import_petro|*",
            "in|final_material|*|export_petro|*",
            "out|secondary|lightoil|loil_imp|*",
            "out|secondary|fueloil|foil_imp|*",
            "out|tertiary_material|clinker_cement|*",
            "out|product|cement|*",
            "out|final_material|BTX|*",
            "out|final_material|ethylene|*",
            "out|final_material|propylene|*",
            "out|secondary|fueloil|agg_ref|*",
            "out|secondary|lightoil|agg_ref|*",
            'out|useful|i_therm|solar_i|M1',
            'out|useful_steel|lt_heat|solar_steel|*',
            'out|useful_aluminum|lt_heat|solar_aluminum|*',
            'out|useful_cement|lt_heat|solar_cement|*',
            'out|useful_petro|lt_heat|solar_petro|*',
            'out|useful_resins|lt_heat|solar_resins|*',
            "in|final|*",
            "in|secondary|coal|coal_NH3|M1",
            "in|secondary|electr|NH3_to_N_fertil|M1",
            'in|secondary|electr|coal_NH3|M1',
            'in|secondary|electr|electr_NH3|M1',
            'in|secondary|electr|gas_NH3|M1',
            'in|secondary|fueloil|fueloil_NH3|M1',
            'in|secondary|gas|gas_NH3|M1',
            "in|secondary|coal|coal_NH3_ccs|M1",
            'in|secondary|electr|coal_NH3_ccs|M1',
            'in|secondary|electr|gas_NH3_ccs|M1',
            'in|secondary|fueloil|fueloil_NH3_ccs|M1',
            'in|secondary|gas|gas_NH3_ccs|M1',
            "in|primary|biomass|biomass_NH3|M1",
            "in|seconday|electr|biomass_NH3|M1",
            "in|primary|biomass|biomass_NH3_ccs|M1",
            "in|secondary|electr|biomass_NH3|M1",
            "in|secondary|electr|biomass_NH3_ccs|M1",
            "in|secondary|electr|fueloil_NH3|M1",
            "in|secondary|electr|fueloil_NH3_ccs|M1",
            "in|secondary|coal|meth_coal|feedstock",
            'in|secondary|coal|meth_coal_ccs|feedstock',
            'in|secondary|gas|meth_ng_ccs|feedstock',
            "in|secondary|gas|meth_ng|feedstock",
            'in|secondary|electr|meth_ng|feedstock',
            'in|secondary|electr|meth_ng_ccs|feedstock',
            "in|secondary|electr|meth_coal|feedstock",
            'in|secondary|electr|meth_coal_ccs|feedstock',
            'in|primary|biomass|meth_bio|feedstock',
            'in|primary|biomass|meth_bio_ccs|feedstock',
            'in|secondary|hydrogen|meth_h2|feedstock',
            'in|secondary|electr|meth_h2|feedstock',
            "in|desulfurized|*|steam_cracker_petro|*",
            "in|secondary_material|*|steam_cracker_petro|*",
            "in|dummy_end_of_life_1|aluminum|scrap_recovery_aluminum_1|M1",
            "in|dummy_end_of_life_2|aluminum|scrap_recovery_aluminum_2|M1",
            "in|dummy_end_of_life_3|aluminum|scrap_recovery_aluminum_3|M1",
            "in|dummy_end_of_life|steel|scrap_recovery_steel|M1",
            "out|dummy_end_of_life_1|aluminum|total_EOL_aluminum|M1",
            "out|dummy_end_of_life_2|aluminum|total_EOL_aluminum|M1",
            "out|dummy_end_of_life_3|aluminum|total_EOL_aluminum|M1",
            "out|dummy_end_of_life|steel|total_EOL_steel|M1",
            "out|dummy_end_of_life|cement|total_EOL_cement|M1",
            "in|product|steel|scrap_recovery_steel|M1",
            "in|final_material|methanol|MTO_petro|M1",
            "in|final_material|methanol|CH2O_synth|M1",
            "out|end_of_life|*",
            "in|end_of_life|*",
            "emis|CO2_industry|*",
            "emis|CF4|*",
            "emis|SO2|*",
            "emis|NOx|*",
            "emis|CH4|*",
            "emis|N2O|*",
            "emis|BCA|*",
            "emis|CO|*",
            "emis|NH3|*",
            "emis|NOx|*",
            "emis|OCA|*",
            "emis|CO2|*",
        ],
        inplace=True,
    )

    # Methanol input conversion from material to energy unit

    df.divide("in|final_material|methanol|MTO_petro|M1", (1/0.6976),
    "in|final_material|methanol|MTO_petro|energy", append=True, ignore_units=True)

    df.divide("in|final_material|methanol|CH2O_synth|M1", (1/0.6976),
    "in|final_material|methanol|CH2O_synth|energy", append=True, ignore_units=True)

    # Convert methanol at primary_material from energy to material unit
    # In the model this is kept as energy units to easily seperate two modes: fuel and feedstock
    df.divide("out|primary_material|methanol|meth_coal|feedstock", 0.6976,
    "out|primary_material|methanol|meth_coal|feedstockMt", append=True, ignore_units=True)

    df.divide("out|primary_material|methanol|meth_coal_ccs|feedstock", 0.6976,
    "out|primary_material|methanol|meth_coal_ccs|feedstockMt", append=True, ignore_units=True)

    df.divide("out|primary_material|methanol|meth_ng|feedstock", 0.6976,
    "out|primary_material|methanol|meth_ng|feedstockMt", append=True, ignore_units=True)

    df.divide("out|primary_material|methanol|meth_ng_ccs|feedstock", 0.6976,
    "out|primary_material|methanol|meth_ng_ccs|feedstockMt", append=True, ignore_units=True)

    df.divide("out|primary_material|methanol|meth_bio|feedstock", 0.6976,
    "out|primary_material|methanol|meth_bio|feedstockMt", append=True, ignore_units=True)

    df.divide("out|primary_material|methanol|meth_bio_ccs|feedstock", 0.6976,
    "out|primary_material|methanol|meth_bio_ccs|feedstockMt", append=True, ignore_units=True)

    df.divide("out|primary_material|methanol|meth_h2|feedstock", 0.6976,
    "out|primary_material|methanol|meth_h2|feedstockMt", append=True, ignore_units=True)

    # Convert methanol at primary from energy to material unit
    # In the model this is kept as energy units to easily seperate two modes: fuel and feedstock
    df.divide("out|primary|methanol|meth_coal|fuel", 0.6976,
    "out|primary|methanol|meth_coal|fuelMt", append=True, ignore_units=True)

    df.divide("out|primary|methanol|meth_coal_ccs|fuel", 0.6976,
    "out|primary|methanol|meth_coal_ccs|fuelMt", append=True, ignore_units=True)

    df.divide("out|primary|methanol|meth_ng|fuel", 0.6976,
    "out|primary|methanol|meth_ng|fuelMt", append=True, ignore_units=True)

    df.divide("out|primary|methanol|meth_ng_ccs|fuel", 0.6976,
    "out|primary|methanol|meth_ng_ccs|fuelMt", append=True, ignore_units=True)

    df.divide("out|primary|methanol|meth_bio|fuel", 0.6976,
    "out|primary|methanol|meth_bio|fuelMt", append=True, ignore_units=True)

    df.divide("out|primary|methanol|meth_bio_ccs|fuel", 0.6976,
    "out|primary|methanol|meth_bio_ccs|fuelMt", append=True, ignore_units=True)

    df.divide("out|primary|methanol|meth_h2|fuel", 0.6976,
    "out|primary|methanol|meth_h2|fuelMt", append=True, ignore_units=True)

    df.convert_unit('unknown', to='', factor=1, inplace = True)

    variables = df.variable
    df.aggregate_region(variables, region="World", method=sum, append=True)

    name = os.path.join(directory, "check.xlsx")
    df.to_excel(name)
    print("Necessary variables are filtered")

    # Obtain the model and scenario name
    model_name = df.model[0]
    scenario_name = df.scenario[0]

    # Create an empty pyam dataframe to store the new variables

    workbook = xlsxwriter.Workbook("empty_template.xlsx")
    worksheet = workbook.add_worksheet()
    worksheet.write("A1", "Model")
    worksheet.write("B1", "Scenario")
    worksheet.write("C1", "Region")
    worksheet.write("D1", "Variable")
    worksheet.write("E1", "Unit")
    columns = [
        "F1",
        "G1",
        "H1",
        "I1",
        "J1",
        "K1",
        "L1",
        "M1",
        "N1",
        "O1",
        "P1",
        "Q1",
        "R1",
    ]

    for yr, col in zip(years, columns):
        worksheet.write(col, yr)
    workbook.close()

    df_final = pyam.IamDataFrame("empty_template.xlsx")
    print("Empty template for new variables created")

    # Create a pdf file with figures
    path = os.path.join(directory, "Material_global_graphs.pdf")
    pp = PdfPages(path)
    # pp = PdfPages("Material_global_graphs.pdf")

    # Reporting and Plotting

    print("Production plots and variables are being generated")
    for r in nodes:
        ## PRODUCTION - PLOTS
        ## Needs to be checked again to see whether the graphs are correct

        fig, (ax1, ax2, ax3) = plt.subplots(1, 3, figsize=(20, 10))
        fig.tight_layout(pad=10.0)

        # ALUMINUM
        df_al = df.copy()
        df_al.filter(region=r, year=years, inplace=True)
        df_al.filter(variable=["out|*|aluminum|*", "in|*|aluminum|*"], inplace=True)
        df_al.convert_unit('', to='Mt/yr', factor=1, inplace = True)
        df_al_graph = df_al.copy()

        df_al_graph.filter(
            variable=[
                "out|useful_material|aluminum|import_aluminum|*",
                "out|final_material|aluminum|prebake_aluminum|*",
                "out|final_material|aluminum|soderberg_aluminum|*",
                "out|new_scrap|aluminum|*",
            ],
            inplace=True,
        )

        if r == "World":
            df_al_graph.filter(
                variable=[
                    "out|final_material|aluminum|prebake_aluminum|*",
                    "out|final_material|aluminum|soderberg_aluminum|*",
                    "out|new_scrap|aluminum|*",
                ],
                inplace=True,
            )

        df_al_graph.plot.stack(ax=ax1)
        ax1.legend(
            [
                "Prebake",
                "Soderberg",
                "Newscrap",
                "Oldscrap_min",
                "Oldscrap_av",
                "Oldscrap_max",
                "import",
            ],
            bbox_to_anchor=(-0.4, 1),
            loc="upper left",
        )
        ax1.set_title("Aluminium Production_" + r)
        ax1.set_xlabel("Year")
        ax1.set_ylabel("Mt")

        # STEEL

        df_steel = df.copy()
        df_steel.filter(region=r, year=years, inplace=True)
        df_steel.filter(variable=["out|*|steel|*", "in|*|steel|*"], inplace=True)
        df_steel.convert_unit('', to='Mt/yr', factor=1, inplace = True)

        df_steel_graph = df_steel.copy()
        df_steel_graph.filter(
            variable=[
                "out|final_material|steel|*",
                "out|useful_material|steel|import_steel|*",
            ],
            inplace=True,
        )

        if r == "World":
            df_steel.filter(variable=["out|*|steel|*", "in|*|steel|*"], inplace=True)
            df_steel_graph.filter(
                variable=["out|final_material|steel|*",], inplace=True,
            )

        df_steel_graph.plot.stack(ax=ax2)
        ax2.legend(
            ["Bof steel", "Eaf steel M1", "Eaf steel M2", "Import"],
            bbox_to_anchor=(-0.4, 1),
            loc="upper left",
        )
        ax2.set_title("Steel Production_" + r)
        ax2.set_ylabel("Mt")

        # PETRO

        df_petro = df.copy()
        df_petro.filter(region=r, year=years, inplace=True)
        df_petro.filter(
            variable=[
                "in|final|ethanol|ethanol_to_ethylene_petro|M1",
                "in|desulfurized|*|steam_cracker_petro|*",
                "in|secondary_material|*|steam_cracker_petro|*",
            ],
            inplace=True,
        )
        df_petro.convert_unit('', to='Mt/yr', factor=1, inplace = True)

        if r == "World":

            df_petro.filter(
                variable=[
                    "in|final|ethanol|ethanol_to_ethylene_petro|M1",
                    "in|desulfurized|*|steam_cracker_petro|*",
                    "in|secondary_material|*|steam_cracker_petro|*",
                ],
                inplace=True,
            )

        df_petro.plot.stack(ax=ax3)
        ax3.legend(
            [
                "atm_gasoil",
                "naphtha",
                "vacuum_gasoil",
                "bioethanol",
                "ethane",
                "propane",
            ],
            bbox_to_anchor=(-0.4, 1),
            loc="upper left",
        )
        ax3.set_title("HVC feedstock" + r)
        ax3.set_xlabel("Years")
        ax3.set_ylabel("GWa")

        plt.close()
        pp.savefig(fig)

        # PRODUCTION - IAMC Variables

        # ALUMINUM

        # Primary Production
        primary_al_vars = [
            "out|final_material|aluminum|prebake_aluminum|M1",
            "out|final_material|aluminum|soderberg_aluminum|M1",
        ]

        # Secondary Production
        secondary_al_vars = ["out|final_material|aluminum|secondary_aluminum|M1"]

        # Collected Scrap
        collected_scrap_al_vars = [
            "out|new_scrap|aluminum|manuf_aluminum|M1",
            "in|dummy_end_of_life_1|aluminum|scrap_recovery_aluminum_1|M1",
            "in|dummy_end_of_life_2|aluminum|scrap_recovery_aluminum_2|M1",
            "in|dummy_end_of_life_3|aluminum|scrap_recovery_aluminum_3|M1"
        ]

        # Total Available Scrap:
        #  New scrap + The end of life products (exegenous assumption)
        # + from power and buildings sector

        new_scrap_al_vars = ["out|new_scrap|aluminum|manuf_aluminum|M1"]
        old_scrap_al_vars = ["out|dummy_end_of_life_1|aluminum|total_EOL_aluminum|M1",
                             "out|dummy_end_of_life_2|aluminum|total_EOL_aluminum|M1",
                             "out|dummy_end_of_life_3|aluminum|total_EOL_aluminum|M1"]

        df_al.aggregate(
            "Production|Primary|Non-Ferrous Metals|Aluminium",
            components=primary_al_vars,
            append=True,
        )
        df_al.aggregate(
            "Production|Secondary|Non-Ferrous Metals|Aluminium",
            components=secondary_al_vars,
            append=True,
        )

        df_al.aggregate(
            "Production|Non-Ferrous Metals|Aluminium",
            components=secondary_al_vars + primary_al_vars,
            append=True,
        )

        df_al.aggregate(
            "Collected Scrap|Non-Ferrous Metals|Aluminium",
            components=collected_scrap_al_vars,
            append=True,
        )

        df_al.aggregate(
            "Collected Scrap|Non-Ferrous Metals",
            components=collected_scrap_al_vars,
            append=True,
        )

        df_al.aggregate(
            "Total Scrap|Non-Ferrous Metals|Aluminium",
            components=new_scrap_al_vars+old_scrap_al_vars,
            append=True,
        )

        df_al.aggregate(
            "Total Scrap|Non-Ferrous Metals",
            components=new_scrap_al_vars+old_scrap_al_vars,
            append=True,
        )

        df_al.aggregate(
            "Total Scrap|Non-Ferrous Metals|Aluminium|New Scrap",
            components=new_scrap_al_vars,
            append=True,
        )

        df_al.aggregate(
            "Total Scrap|Non-Ferrous Metals|Aluminium|Old Scrap",
            components=old_scrap_al_vars,
            append=True,
        )

        df_al.aggregate(
            "Production|Non-Ferrous Metals",
            components=primary_al_vars + secondary_al_vars,
            append=True,
        )

        df_al.filter(
            variable=[
                "Production|Primary|Non-Ferrous Metals|Aluminium",
                "Production|Secondary|Non-Ferrous Metals|Aluminium",
                "Production|Non-Ferrous Metals|Aluminium",
                "Production|Non-Ferrous Metals",
                "Collected Scrap|Non-Ferrous Metals|Aluminium",
                "Collected Scrap|Non-Ferrous Metals",
                "Total Scrap|Non-Ferrous Metals",
                "Total Scrap|Non-Ferrous Metals|Aluminium",
                "Total Scrap|Non-Ferrous Metals|Aluminium|New Scrap",
                "Total Scrap|Non-Ferrous Metals|Aluminium|Old Scrap",
            ],
            inplace=True,
        )

        # STEEL

        primary_steel_vars = ["out|final_material|steel|bof_steel|M1",
                              "out|final_material|steel|eaf_steel|M1",
                              "out|final_material|steel|eaf_steel|M3"
                              ]

        secondary_steel_vars = [
            "out|final_material|steel|eaf_steel|M2",
            "in|new_scrap|steel|bof_steel|M1"
        ]

        collected_scrap_steel_vars = [
            "in|dummy_end_of_life|steel|scrap_recovery_steel|M1"
        ]
        total_scrap_steel_vars = ["out|dummy_end_of_life|steel|total_EOL_steel|M1"]

        new_scrap_steel_vars = ["out|new_scrap|steel|manuf_steel|M1"]
        old_scrap_steel_vars = ["out|dummy_end_of_life|steel|total_EOL_steel|M1"]

        df_steel.aggregate(
            "Production|Primary|Steel (before sub.)", components=primary_steel_vars, append=True,
        )

        df_steel.subtract("Production|Primary|Steel (before sub.)",
        "in|new_scrap|steel|bof_steel|M1","Production|Primary|Steel", append = True)

        df_steel.aggregate(
            "Production|Secondary|Steel", components=secondary_steel_vars, append=True,
        )

        df_steel.aggregate(
            "Production|Steel",
            components=["Production|Primary|Steel", "Production|Secondary|Steel"],
            append=True,
        )

        df_steel.aggregate(
            "Collected Scrap|Steel", components=collected_scrap_steel_vars, append=True,
        )
        df_steel.aggregate(
            "Total Scrap|Steel", components=total_scrap_steel_vars, append=True
        )

        df_steel.aggregate(
            "Total Scrap|Steel|Old Scrap", components=old_scrap_steel_vars, append=True
        )

        #df_steel.aggregate(
        #    "Total Scrap|Steel|New Scrap", components=new_scrap_steel_vars, append=True
        #)

        df_steel.filter(
            variable=[
                "Production|Primary|Steel",
                "Production|Secondary|Steel",
                "Production|Steel",
                "Collected Scrap|Steel",
                "Total Scrap|Steel",
                "Total Scrap|Steel|Old Scrap",
                "Total Scrap|Steel|New Scrap",
            ],
            inplace=True,
        )

        # CHEMICALS

        df_chemicals = df.copy()
        df_chemicals.filter(region=r, year=years, inplace=True)
        df_chemicals.filter(variable=['out|secondary_material|NH3|*',
                                      "out|final_material|ethylene|*",
                                      "out|final_material|propylene|*",
                                       "out|final_material|BTX|*",
                                       'out|primary_material|methanol|*|feedstockMt',
                                       'out|primary|methanol|*|fuelMt'
                                       ],inplace=True)
        df_chemicals.convert_unit('', to='Mt/yr', factor=1, inplace = True)
        df_chemicals.convert_unit('GWa', to='Mt/yr', factor=(1/0.6976), inplace=True)

       # Methanol

       # In Mt units
        primary_methanol_chemical_vars = ["out|primary_material|methanol|meth_coal|feedstockMt",
                                          "out|primary_material|methanol|meth_coal_ccs|feedstockMt",
                                          "out|primary_material|methanol|meth_ng|feedstockMt",
                                          "out|primary_material|methanol|meth_ng_ccs|feedstockMt",
                                          "out|primary_material|methanol|meth_bio|feedstockMt",
                                          "out|primary_material|methanol|meth_bio_ccs|feedstockMt",
                                          "out|primary_material|methanol|meth_h2|feedstockMt",
                                        ]
        methanol_fuel_vars = ["out|primary|methanol|meth_coal|fuelMt",
                              "out|primary|methanol|meth_coal_ccs|fuelMt",
                              "out|primary|methanol|meth_ng|fuelMt",
                              "out|primary|methanol|meth_ng_ccs|fuelMt",
                              "out|primary|methanol|meth_bio|fuelMt",
                              "out|primary|methanol|meth_bio_ccs|fuelMt",
                              "out|primary|methanol|meth_h2|fuelMt",
                             ]

        methanol_total_vars = primary_methanol_chemical_vars + methanol_fuel_vars

        df_chemicals.aggregate(
            "Production|Methanol",
            components=methanol_total_vars,
            append=True,
        )

        df_chemicals.aggregate(
            "Production|Chemicals|Methanol",
            components=primary_methanol_chemical_vars,
            append=True,
        )

        df_chemicals.aggregate(
            "Production|Primary|Chemicals|Methanol",
            components=primary_methanol_chemical_vars,
            append=True,
        )

        df_chemicals.aggregate(
            "Production|Fuel|Methanol",
            components=methanol_fuel_vars,
            append=True,
        )

        # add entries for each methanol technology
        meth_tec_list = [i.replace("fuel", "M1") for i in methanol_fuel_vars]
        df_meth_individual = df_chemicals.filter(variable=meth_tec_list)
        df_meth_individual.convert_unit('Mt/yr', to='Mt/yr', factor=(1/0.6976), inplace=True)
        var_name = "Production|Methanol|"
        for i in df_meth_individual["variable"]:
            df_meth_individual.rename({"variable": {i: i.replace("out|primary|methanol|", var_name).replace("|M1", "")}},
            inplace=True)

        # AMMONIA

        primary_ammonia_vars = [
            "out|secondary_material|NH3|gas_NH3|M1",
            "out|secondary_material|NH3|gas_NH3_ccs|M1",
            "out|secondary_material|NH3|coal_NH3|M1",
            "out|secondary_material|NH3|coal_NH3_ccs|M1",
            "out|secondary_material|NH3|biomass_NH3|M1",
            "out|secondary_material|NH3|biomass_NH3_ccs|M1",
            "out|secondary_material|NH3|fueloil_NH3|M1",
            "out|secondary_material|NH3|fueloil_NH3_ccs|M1",
            "out|secondary_material|NH3|electr_NH3|M1"
        ]

        df_chemicals.aggregate(
            "Production|Primary|Chemicals|Ammonia",
            components=primary_ammonia_vars,
            append=True,
        )

        df_chemicals.aggregate(
            "Production|Chemicals|Ammonia",
            components=primary_ammonia_vars,
            append=True,
        )

       # High Value Chemicals

        intermediate_petro_vars = [
            "out|final_material|ethylene|ethanol_to_ethylene_petro|M1",
            "out|final_material|ethylene|steam_cracker_petro|atm_gasoil",
            "out|final_material|ethylene|steam_cracker_petro|naphtha",
            "out|final_material|ethylene|steam_cracker_petro|vacuum_gasoil",
            "out|final_material|ethylene|steam_cracker_petro|ethane",
            "out|final_material|ethylene|steam_cracker_petro|propane",
            "out|final_material|propylene|steam_cracker_petro|atm_gasoil",
            "out|final_material|propylene|steam_cracker_petro|naphtha",
            "out|final_material|propylene|steam_cracker_petro|vacuum_gasoil",
            "out|final_material|propylene|steam_cracker_petro|propane",
            "out|final_material|BTX|steam_cracker_petro|atm_gasoil",
            "out|final_material|BTX|steam_cracker_petro|naphtha",
            "out|final_material|BTX|steam_cracker_petro|vacuum_gasoil",
            "out|final_material|ethylene|MTO_petro|M1",
            "out|final_material|propylene|MTO_petro|M1",
        ]

        df_chemicals.aggregate(
            "Production|Primary|Chemicals|High Value Chemicals",
            components=intermediate_petro_vars,
            append=True,
        )

        df_chemicals.aggregate(
            "Production|Chemicals|High Value Chemicals",
            components=intermediate_petro_vars,
            append=True,
        )

        # Totals

        chemicals_vars = intermediate_petro_vars + primary_ammonia_vars + primary_methanol_chemical_vars
        df_chemicals.aggregate(
            "Production|Primary|Chemicals",
            components=chemicals_vars,
            append=True,
        )

        df_chemicals.aggregate(
            "Production|Chemicals",
            components=chemicals_vars,
            append=True,
        )

        df_chemicals.filter(
            variable=[
                "Production|Primary|Chemicals|High Value Chemicals",
                "Production|Chemicals|High Value Chemicals",
                "Production|Primary|Chemicals",
                "Production|Chemicals",
                "Production|Primary|Chemicals|Ammonia",
                "Production|Chemicals|Ammonia",
                "Production|Primary|Chemicals|Methanol",
                "Production|Chemicals|Methanol",
                "Production|Fuel|Methanol",
                'Production|Methanol'
            ],
            inplace=True,
        )

        df_chemicals.append(df_meth_individual, inplace=True)

        # Add to final data_frame
        df_final.append(df_al, inplace=True)
        df_final.append(df_steel, inplace=True)
        df_final.append(df_chemicals, inplace=True)

    # CEMENT

    for r in nodes:

        # PRODUCTION - PLOT

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(15, 10))
        fig.tight_layout(pad=15.0)

        # Clinker cement

        df_cement_clinker = df.copy()
        df_cement_clinker.filter(region=r, year=years, inplace=True)
        df_cement_clinker.filter(
            variable=["out|tertiary_material|clinker_cement|*"], inplace=True
        )
        df_cement_clinker.plot.stack(ax=ax1)
        ax1.legend(
            ["Dry Clinker", "Wet Clinker"], bbox_to_anchor=(-0.5, 1), loc="upper left"
        )
        ax1.set_title("Clinker Cement Production_" + r)
        ax1.set_xlabel("Year")
        ax1.set_ylabel("Mt")

        # Final prodcut cement

        df_cement = df.copy()
        df_cement.filter(region=r, year=years, inplace=True)
        df_cement.filter(variable=["out|product|cement|*",
                                   "out|tertiary_material|clinker_cement|*"
                                  ], inplace=True)
        # df_cement.plot.stack(ax=ax2)
        # ax2.legend(
        #     ["Ballmill Grinding", "Vertical Mill Grinding"],
        #     bbox_to_anchor=(-0.6, 1),
        #     loc="upper left",
        # )
        # ax2.set_title("Final Cement Production_" + r)
        # ax2.set_xlabel("Year")
        # ax2.set_ylabel("Mt")

        plt.close()
        pp.savefig(fig)

        # PRODUCTION - IAMC format

        primary_cement_vars = [
            "out|product|cement|grinding_ballmill_cement|M1",
            "out|product|cement|grinding_vertmill_cement|M1",
        ]

        clinker_vars = [
        "out|tertiary_material|clinker_cement|clinker_dry_cement|M1",
        "out|tertiary_material|clinker_cement|clinker_wet_cement|M1"
        ]

        total_scrap_cement_vars = ["out|dummy_end_of_life|cement|total_EOL_cement|M1"]

        df_cement.convert_unit('', to='Mt/yr', factor=1, inplace = True)

        df_cement.aggregate(
            "Production|Non-Metallic Minerals|Clinker", components=clinker_vars, append=True,
        )

        df_cement.aggregate(
            "Production|Primary|Non-Metallic Minerals|Cement", components=primary_cement_vars, append=True,
        )

        df_cement.aggregate(
            "Production|Non-Metallic Minerals",
            components=primary_cement_vars,
            append=True,
        )

        df_cement.aggregate(
            "Production|Non-Metallic Minerals|Cement", components=primary_cement_vars, append=True,
        )

        df_cement.aggregate(
            "Total Scrap|Non-Metallic Minerals",
            components=total_scrap_cement_vars,
            append=True,
        )

        df_cement.aggregate(
            "Total Scrap|Non-Metallic Minerals|Cement",
            components=total_scrap_cement_vars,
            append=True,
        )
        df_cement.filter(
            variable=[
                "Production|Primary|Non-Metallic Minerals|Cement",
                "Production|Non-Metallic Minerals|Cement",
                "Total Scrap|Non-Metallic Minerals"
                "Total Scrap|Non-Metallic Minerals|Cement",
                "Production|Non-Metallic Minerals|Clinker",
            ],
            inplace=True,
        )
        df_final.append(df_cement, inplace=True)

    # ---------------------------------------------------------------------------------------
    # FINAL ENERGY BY FUELS (Only Non-Energy Use)
    # HVC production, ammonia production and methanol production.

    print("Final Energy by fuels only non-energy use is being printed.")
    commodities = ["gas", "liquids", "solids",'hydrogen','methanol',"all",'electr_gas']

    for c in commodities:
        for r in nodes:
            df_final_energy = df.copy()

            # GWa to EJ/yr
            df_final_energy.convert_unit("", to="GWa", factor=1, inplace=True)
            df_final_energy.convert_unit(
                "GWa", to="EJ/yr", factor=0.03154, inplace=True)
            df_final_energy.filter(region=r, year=years, inplace=True)
            df_final_energy.filter(
                variable=["in|final|*|cokeoven_steel|*"], keep=False, inplace=True
            )
            df_final_energy.filter(
                variable=[
                    "in|final|atm_gasoil|*",
                    "in|final|vacuum_gasoil|*",
                    "in|final|naphtha|*",
                    "in|final|*|coal_fs|*",
                    "in|final|*|methanol_fs|*",
                    "in|final|*|ethanol_fs|*",
                    "in|final|ethanol|ethanol_to_ethylene_petro|*",
                    "in|final|*|foil_fs|*",
                    "in|final|gas|gas_processing_petro|*",
                    "in|final|*|loil_fs|*",
                    "in|final|*|gas_fs|*",
                    'in|secondary|gas|gas_NH3|M1',
                    'in|secondary|gas|gas_NH3_ccs|M1',
                    'in|secondary|fueloil|fueloil_NH3|M1',
                    'in|secondary|fueloil|fueloil_NH3_ccs|M1',
                    'in|secondary|electr|electr_NH3|M1',
                     "in|secondary|coal|coal_NH3|M1",
                     "in|secondary|coal|coal_NH3_ccs|M1",
                     'in|primary|biomass|biomass_NH3_ccs|M1',
                     'in|primary|biomass|biomass_NH3|M1',
                     "in|final_material|methanol|MTO_petro|energy",
                     "in|final_material|methanol|CH2O_synth|energy",
                     'in|secondary|coal|meth_coal|feedstock',
                     'in|secondary|coal|meth_coal_ccs|feedstock',
                     'in|secondary|gas|meth_ng|feedstock',
                     'in|secondary|gas|meth_ng_ccs|feedstock',
                     'in|primary|biomass|meth_bio|feedstock',
                     'in|primary|biomass|meth_bio_ccs|feedstock',
                     'in|secondary|hydrogen|meth_h2|feedstock',
                ],
                inplace=True,
            )

            if c == 'all':
                df_final_energy.filter(variable=["in|final_material|methanol|MTO_petro|energy",
                                     "in|final_material|methanol|CH2O_synth|energy",],
                                                keep=False,inplace=True)
            if c == 'electr_gas':
                df_final_energy.filter(variable=['in|secondary|electr|electr_NH3|M1',],
                                                  inplace=True)
            if c == "gas":
                df_final_energy.filter(variable=["in|final|gas|*",
                                                'in|secondary|gas|gas_NH3|M1',
                                                'in|secondary|gas|gas_NH3_ccs|M1',
                                                'in|secondary|gas|meth_ng|feedstock',
                                                'in|secondary|gas|meth_ng_ccs|feedstock',
                                                'in|secondary|electr|electr_NH3|M1'],
                                                inplace=True)
                df_final_energy.filter(variable=["in|final|gas|gas_processing_petro|*"],
                                                keep=False, inplace=True)
            if c == "liquids":
                df_final_energy.filter(
                    variable=[
                        "in|final|ethanol|ethanol_to_ethylene_petro|*",
                        "in|final|*|foil_fs|*",
                        "in|final|*|loil_fs|*",
                        "in|final|*|methanol_fs|*",
                        "in|final|*|ethanol_fs|*",
                        "in|final|atm_gasoil|*",
                        "in|final|vacuum_gasoil|*",
                        "in|final|naphtha|*",
                        'in|secondary|fueloil|fueloil_NH3|M1',
                        'in|secondary|fueloil|fueloil_NH3_ccs|M1',
                        "in|final|gas|gas_processing_petro|*"
                    ],
                    inplace=True,
                )
            if c == "solids":
                df_final_energy.filter(
                    variable=["in|final|biomass|*", "in|final|coal|*",
                              "in|secondary|coal|coal_NH3|M1",
                              "in|secondary|coal|coal_NH3_ccs|M1",
                              'in|primary|biomass|biomass_NH3_ccs|M1',
                              'in|primary|biomass|biomass_NH3|M1',
                              'in|secondary|coal|meth_coal|feedstock',
                              'in|secondary|coal|meth_coal_ccs|feedstock',
                              'in|primary|biomass|meth_bio|feedstock',
                              'in|primary|biomass|meth_bio_ccs|feedstock',
                              ], inplace=True)
            if c == "hydrogen":
                df_final_energy.filter(
                    variable=[
                              'in|secondary|hydrogen|meth_h2|feedstock',
                              ], inplace=True)
            if c == "methanol":
                df_final_energy.filter(
                    variable=[
                           "in|final_material|methanol|MTO_petro|energy",
                           "in|final_material|methanol|CH2O_synth|energy",
                              ], inplace=True)

            all_flows = df_final_energy.timeseries().reset_index()
            splitted_vars = [v.split("|") for v in all_flows.variable]
            aux1_df = pd.DataFrame(
                splitted_vars,
                columns=["flow_type", "level", "commodity", "technology", "mode"],
            )
            aux2_df = pd.concat(
                [all_flows.reset_index(drop=True), aux1_df.reset_index(drop=True)],
                axis=1,
            )

            # Include only the related industry sector variables
            # Aluminum, cement and steel do not have feedstock use

            var_sectors = [v for v in aux2_df["variable"].values]

            aux2_df = aux2_df[aux2_df["variable"].isin(var_sectors)]

            df_final_energy.filter(variable=var_sectors, inplace=True)

            # Aggregate

            if c == "all":
                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use", components=var_sectors, append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Non-Energy Use"], inplace=True
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

            if c == "methanol":

                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Other",
                    components=var_sectors,
                    append=True,
                )

                df_final_energy.filter(
                    variable=["Final Energy|Non-Energy Use|Other"], inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

            if c == "hydrogen":

                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Hydrogen",
                    components=var_sectors,
                    append=True,
                )

                df_final_energy.filter(
                    variable=["Final Energy|Non-Energy Use|Hydrogen"], inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

            if c == "electr_gas":

                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Gases|Electricity",
                    components=var_sectors,
                    append=True,
                )

                df_final_energy.filter(
                    variable=["Final Energy|Non-Energy Use|Gases|Electricity"], inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

            if c == "gas":

                # Can not distinguish by type Gases (natural gas, biomass, synthetic fossil, efuel)
                # (coal_gas), from biomass (gas_bio), natural gas (gas_bal): All go into secondary level
                # Can not be distinguished in the final level.
                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Gases",
                    components=var_sectors,
                    append=True,
                )

                df_final_energy.filter(
                    variable=["Final Energy|Non-Energy Use|Gases"], inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

            if c == "liquids":

                # All liquids
                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Liquids",
                    components=var_sectors,
                    append=True,
                )

                # Only bios

                filter_vars = [
                    v for v in aux2_df["variable"].values if (("ethanol" in v)
                    & ("methanol" not in v))
                ]
                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Liquids|Biomass",
                    components=filter_vars,
                    append=True,
                )
                # Fossils

                filter_vars = [
                    v
                    for v in aux2_df["variable"].values
                    if (
                        ("lightoil" in v)
                        | ("fueloil" in v)
                        | ("atm_gasoil" in v)
                        | ("vacuum_gasoil" in v)
                        | ("naphtha" in v)
                    )
                ]

                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Liquids|Oil",
                    components=filter_vars,
                    append=True,
                )

                # Natural Gas Liquids (Ethane/Propane)

                filter_vars = [
                    v
                    for v in aux2_df["variable"].values
                    if (
                        ("gas_proc" in v)
                    )
                ]

                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Liquids|Gas",
                    components=filter_vars,
                    append=True,
                )

                df_final_energy.filter(
                    variable=[
                        "Final Energy|Non-Energy Use|Liquids",
                        "Final Energy|Non-Energy Use|Liquids|Oil",
                        "Final Energy|Non-Energy Use|Liquids|Biomass",
                        "Final Energy|Non-Energy Use|Liquids|Gas"
                    ],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "solids":

                # All
                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Solids",
                    components=var_sectors,
                    append=True,
                )
                # Bio
                filter_vars = [
                    v for v in aux2_df["variable"].values if ("biomass" in v)
                ]
                if filter_vars:
                    df_final_energy.aggregate(
                        "Final Energy|Non-Energy Use|Solids|Biomass",
                        components=filter_vars,
                        append=True,
                    )

                # Fossil
                filter_vars = [v for v in aux2_df["variable"].values if ("coal" in v)]
                df_final_energy.aggregate(
                    "Final Energy|Non-Energy Use|Solids|Coal",
                    components=filter_vars,
                    append=True,
                )
                df_final_energy.filter(
                    variable=[
                        "Final Energy|Non-Energy Use|Solids",
                        "Final Energy|Non-Energy Use|Solids|Biomass",
                        "Final Energy|Non-Energy Use|Solids|Coal",
                    ],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

    # FINAL ENERGY BY FUELS (Excluding Non-Energy Use)
    # For ammonia and methanol only electricity use is included since only this
    # has seperate input values in the model.

    print("Final Energy by fuels excluding non-energy use is being printed.")
    commodities = ["electr", "gas", "hydrogen", "liquids", "solids", "heat", 'solar', "all"]
    for c in commodities:

        for r in nodes:

            df_final_energy = df.copy()
            df_final_energy.convert_unit("", to="GWa", factor=1, inplace=True)
            df_final_energy.filter(region=r, year=years, inplace=True)
            df_final_energy.filter(
                variable=["in|final|*|cokeoven_steel|*",
                          "in|final|co_gas|*",
                          "in|final|bf_gas|*"], keep=False, inplace=True
            )

            if c == "electr":
                df_final_energy.filter(variable=["in|final|electr|*",
                                                 'in|secondary|electr|NH3_to_N_fertil|M1',
                                                 'in|secondary|electr|coal_NH3|M1',
                                                 'in|secondary|electr|fueloil_NH3|M1',
                                                 'in|secondary|electr|gas_NH3|M1',
                                                 'in|secondary|electr|coal_NH3_ccs|M1',
                                                 'in|secondary|electr|fueloil_NH3_ccs|M1',
                                                 'in|secondary|electr|gas_NH3_ccs|M1',
                                                 'in|secondary|electr|biomass_NH3_ccs|M1',
                                                 'in|secondary|electr|biomass_NH3|M1',
                                                 'in|secondary|electr|meth_ng|feedstock',
                                                 'in|secondary|electr|meth_ng_ccs|feedstock',
                                                 'in|secondary|electr|meth_coal|feedstock',
                                                 'in|secondary|electr|meth_coal_ccs|feedstock',
                                                 'in|secondary|electr|meth_h2|feedstock'
                                                 ], inplace=True)
            if c == "gas":
                df_final_energy.filter(variable=["in|final|gas|*"], inplace=True)
            # Do not include gasoil and naphtha feedstock
            if c == "liquids":
                df_final_energy.filter(
                    variable=[
                        "in|final|ethanol|*",
                        "in|final|fueloil|*",
                        "in|final|lightoil|*",
                        "in|final|methanol|*",
                    ],
                    inplace=True,
                )
            if c == "solids":
                df_final_energy.filter(
                    variable=[
                        "in|final|biomass|*",
                        "in|final|coal|*",
                        "in|final|coke_iron|*",
                    ],
                    inplace=True,
                )
            if c == "hydrogen":
                df_final_energy.filter(variable=["in|final|hydrogen|*"], inplace=True)
            if c == "heat":
                df_final_energy.filter(variable=["in|final|d_heat|*"], inplace=True)
            if c == 'solar':
                df_final_energy.filter(variable=["out|useful|i_therm|solar_i|*",
                                                 'out|useful_aluminum|lt_heat|solar_aluminum|*',
                                                 'out|useful_steel|lt_heat|solar_steel|*',
                                                 'out|useful_cement|lt_heat|solar_cement|*',
                                                 'out|useful_petro|lt_heat|solar_petro|*',
                                                 'out|useful_resins|lt_heat|solar_resins|*',
                ], inplace=True)
            if c == "all":
                df_final_energy.filter(variable=["in|final|*",
                                                 'in|secondary|electr|NH3_to_N_fertil|M1',
                                                 'in|secondary|electr|coal_NH3|M1',
                                                 'in|secondary|electr|fueloil_NH3|M1',
                                                 'in|secondary|electr|gas_NH3|M1',
                                                 'in|secondary|electr|coal_NH3_ccs|M1',
                                                 'in|secondary|electr|fueloil_NH3_ccs|M1',
                                                 'in|secondary|electr|gas_NH3_ccs|M1',
                                                 'in|secondary|electr|biomass_NH3_ccs|M1',
                                                 'in|secondary|electr|biomass_NH3|M1',
                                                 'out|useful|i_therm|solar_i|M1',
                                                 'out|useful_aluminum|lt_heat|solar_aluminum|*',
                                                 'out|useful_steel|lt_heat|solar_steel|*',
                                                 'out|useful_cement|lt_heat|solar_cement|*',
                                                 'out|useful_petro|lt_heat|solar_petro|*',
                                                 'out|useful_resins|lt_heat|solar_resins|*',
                                                  'in|secondary|electr|meth_ng_ccs|feedstock',
                                                  'in|secondary|electr|meth_ng|feedstock',
                                                  'in|secondary|electr|meth_coal|feedstock',
                                                  'in|secondary|electr|meth_coal_ccs|feedstock',
                                                  'in|secondary|electr|meth_h2|feedstock'
                                                 ], inplace=True)

            all_flows = df_final_energy.timeseries().reset_index()
            splitted_vars = [v.split("|") for v in all_flows.variable]
            aux1_df = pd.DataFrame(
                splitted_vars,
                columns=["flow_type", "level", "commodity", "technology", "mode"],
            )
            aux2_df = pd.concat(
                [all_flows.reset_index(drop=True), aux1_df.reset_index(drop=True)],
                axis=1,
            )

            # Include only the related industry sector variables and state some
            # exceptions
            var_sectors = [
                v for v in aux2_df["variable"].values
                if ((
                          (v.split('|')[3].endswith("cement"))
                         | (v.split('|')[3].endswith("steel"))
                         | (v.split('|')[3].endswith("aluminum"))
                         | (v.split('|')[3].endswith("petro"))
                         | (v.split('|')[3].endswith("resins"))
                         | (v.split('|')[3].endswith("_i"))
                         | (v.split('|')[3].endswith("_I"))
                         | (('NH3') in v)
                         | (v.split('|')[3].startswith("meth"))
                         | (v.split('|')[3].startswith("CH2O"))

                         )
                        & (
                                ("ethanol_to_ethylene_petro" not in v)
                                & ("gas_processing_petro" not in v)
                                & (
                                        "in|final|atm_gasoil|steam_cracker_petro|atm_gasoil"
                                        not in v
                                )
                                & (
                                        "in|final|vacuum_gasoil|steam_cracker_petro|vacuum_gasoil"
                                        not in v
                                )
                                & ("in|final|naphtha|steam_cracker_petro|naphtha" not in v)
                        ))
            ]
            aux2_df = aux2_df[aux2_df["variable"].isin(var_sectors)]

            df_final_energy.filter(variable=var_sectors, inplace=True)

            # Aggregate

            if c == "all":
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use",
                    components=var_sectors,
                    append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Industry excl Non-Energy Use"], inplace=True
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "electr":
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Electricity",
                    components=var_sectors,
                    append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Industry excl Non-Energy Use|Electricity"],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "gas":

                # Can not distinguish by type Gases (natural gas, biomass, synthetic fossil, efuel)
                # (coal_gas), from biomass (gas_bio), natural gas (gas_bal): All go into secondary level
                # Can not be distinguished in the final level.

                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Gases",
                    components=var_sectors,
                    append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Industry excl Non-Energy Use|Gases"],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

            if c == "hydrogen":
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Hydrogen",
                    components=var_sectors,
                    append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Industry excl Non-Energy Use|Hydrogen"],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

            if c == "liquids":

                # All liquids
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Liquids",
                    components=var_sectors,
                    append=True,
                )

                # Only bios (ethanol, methanol ?)

                filter_vars = [
                    v for v in aux2_df["variable"].values if (("ethanol" in v)
                                                              & ('methanol' not in v))
                ]
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Liquids|Biomass",
                    components=filter_vars,
                    append=True,
                )

                # Fossils

                filter_vars = [
                    v for v in aux2_df["variable"].values if (("fueloil" in v)
                                                              | ("lightoil" in v))
                ]

                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Liquids|Oil",
                    components=filter_vars,
                    append=True,
                )

                # Other

                filter_vars = [
                    v for v in aux2_df["variable"].values if (("methanol" in v))
                ]

                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Liquids|Coal",
                    components=filter_vars,
                    append=True,
                )

                df_final_energy.filter(
                    variable=[
                        "Final Energy|Industry excl Non-Energy Use|Liquids",
                        "Final Energy|Industry excl Non-Energy Use|Liquids|Oil",
                        "Final Energy|Industry excl Non-Energy Use|Liquids|Biomass",
                        "Final Energy|Industry excl Non-Energy Use|Liquids|Coal",
                    ],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "solids":

                # All
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Solids",
                    components=var_sectors,
                    append=True,
                )

                # Bio
                filter_vars = [
                    v for v in aux2_df["variable"].values if ("biomass" in v)
                ]
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Solids|Biomass",
                    components=filter_vars,
                    append=True,
                )

                # Fossil
                filter_vars = [
                    v for v in aux2_df["variable"].values if ("biomass" not in v)
                ]
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Solids|Coal",
                    components=filter_vars,
                    append=True,
                )
                df_final_energy.filter(
                    variable=[
                        "Final Energy|Industry excl Non-Energy Use|Solids",
                        "Final Energy|Industry excl Non-Energy Use|Solids|Biomass",
                        "Final Energy|Industry excl Non-Energy Use|Solids|Coal",
                    ],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "heat":
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Heat",
                    components=var_sectors,
                    append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Industry excl Non-Energy Use|Heat"],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == 'solar':
                df_final_energy.aggregate(
                    "Final Energy|Industry excl Non-Energy Use|Solar",
                    components=var_sectors,
                    append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Industry excl Non-Energy Use|Solar"],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

    # FINAL ENERGY BY FUELS (Including Non-Energy Use)
    print("Final Energy by fuels including non-energy use is being printed.")
    commodities = ["electr", "gas", "hydrogen", "liquids", "solids", "heat", "all", 'solar']
    for c in commodities:

        for r in nodes:

            df_final_energy = df.copy()
            df_final_energy.convert_unit(
                "", to="GWa", factor=1, inplace=True)
            df_final_energy.filter(region=r, year=years, inplace=True)

            exclude = [
                "in|final|*|cokeoven_steel|*",
                "in|final|bf_gas|*",
                "in|final|co_gas|*",
                'in|final|*|meth_fc_trp|*',
                'in|final|*|meth_ic_trp|*',
                'in|final|*|meth_i|*',
                'in|final|*|meth_rc|*',
                'in|final|*|sp_meth_I|*']

            df_final_energy.filter(variable=exclude, keep=False, inplace=True)

            if c == 'solar':
                df_final_energy.filter(variable=["out|useful|i_therm|solar_i|M1",
                                                 "out|useful_steel|lt_heat|solar_steel|*",
                                                 "out|useful_aluminum|lt_heat|solar_aluminum|*",
                                                 "out|useful_cement|lt_heat|solar_cement|*",
                                                 "out|useful_petro|lt_heat|solar_petro|*",
                                                 "out|useful_resins|lt_heat|solar_resins|*",
                ],
                                                    inplace = True)
            if c == "electr":
                df_final_energy.filter(variable=["in|final|electr|*",
                                                 'in|secondary|electr|NH3_to_N_fertil|M1',
                                                 'in|secondary|electr|coal_NH3|M1',
                                                 'in|secondary|electr|electr_NH3|M1',
                                                 'in|secondary|electr|fueloil_NH3|M1',
                                                 'in|secondary|electr|gas_NH3|M1',
                                                 'in|secondary|electr|coal_NH3_ccs|M1',
                                                 'in|secondary|electr|fueloil_NH3_ccs|M1',
                                                 'in|secondary|electr|gas_NH3_ccs|M1',
                                                 'in|secondary|electr|biomass_NH3_ccs|M1',
                                                 'in|secondary|electr|biomass_NH3|M1',
                                                 'in|secondary|electr|meth_ng|feedstock',
                                                 'in|secondary|electr|meth_ng_ccs|feedstock',
                                                 'in|secondary|electr|meth_coal|feedstock',
                                                 'in|secondary|electr|meth_coal_ccs|feedstock',
                                                 'in|secondary|electr|meth_h2|feedstock'
                                                 ], inplace=True)
            if c == "gas":
                df_final_energy.filter(variable=["in|final|gas|*",
                                                 'in|secondary|gas|gas_NH3|M1',
                                                 'in|secondary|gas|gas_NH3_ccs|M1',
                                                 'in|secondary|gas|meth_ng_ccs|feedstock',
                                                 "in|secondary|gas|meth_ng|feedstock",
                                                 ],
                                       inplace=True)
                df_final_energy.filter(variable=["in|final|gas|gas_processing_petro|*"],
                                                keep=False, inplace=True)
            # Include gasoil and naphtha feedstock
            if c == "liquids":
                df_final_energy.filter(
                    variable=[
                        "in|final|ethanol|*",
                        "in|final|fueloil|*",
                        "in|final|lightoil|*",
                        "in|final|methanol|*",
                        "in|final|vacuum_gasoil|*",
                        "in|final|naphtha|*",
                        "in|final|atm_gasoil|*",
                        'in|secondary|fueloil|fueloil_NH3|M1',
                        'in|secondary|fueloil|fueloil_NH3_ccs|M1',
                        "in|final|gas|gas_processing_petro|*"],
                    inplace=True,
                )
            if c == "solids":
                df_final_energy.filter(
                    variable=[
                        "in|final|biomass|*",
                        "in|final|coal|*",
                        "in|final|coke_iron|*",
                        'in|secondary|coal|coal_NH3|M1',
                        'in|secondary|coal|coal_NH3_ccs|M1',
                        "in|secondary|coal|meth_coal|feedstock",
                        'in|secondary|coal|meth_coal_ccs|feedstock',
                        'in|primary|biomass|meth_bio|feedstock',
                        'in|primary|biomass|meth_bio_ccs|feedstock',
                        'in|primary|biomass|biomass_NH3_ccs|M1',
                        'in|primary|biomass|biomass_NH3|M1'
                    ],
                    inplace=True,
                )
            if c == "hydrogen":
                df_final_energy.filter(variable=["in|final|hydrogen|*",
                                                 'in|secondary|hydrogen|meth_h2|feedstock'], inplace=True)
            if c == "heat":
                df_final_energy.filter(variable=["in|final|d_heat|*"], inplace=True)
            if c == "all":
                df_final_energy.filter(variable=["in|final|*",
                                                 "in|secondary|coal|coal_NH3|M1",
                                                 "in|secondary|electr|NH3_to_N_fertil|M1",
                                                 'in|secondary|electr|coal_NH3|M1',
                                                 'in|secondary|electr|electr_NH3|M1',
                                                 'in|secondary|electr|gas_NH3|M1',
                                                 'in|secondary|fueloil|fueloil_NH3|M1',
                                                 'in|secondary|electr|fueloil_NH3|M1',
                                                 'in|secondary|gas|gas_NH3|M1',
                                                 "in|secondary|coal|coal_NH3_ccs|M1",
                                                 'in|secondary|electr|coal_NH3_ccs|M1',
                                                 'in|secondary|electr|gas_NH3_ccs|M1',
                                                 'in|secondary|fueloil|fueloil_NH3_ccs|M1',
                                                 'in|secondary|electr|fueloil_NH3_ccs|M1',
                                                 'in|secondary|gas|gas_NH3_ccs|M1',
                                                 "in|primary|biomass|biomass_NH3|M1",
                                                 "in|secondary|electr|biomass_NH3|M1",
                                                 "in|primary|biomass|biomass_NH3_ccs|M1",
                                                 "in|secondary|electr|biomass_NH3_ccs|M1",
                                                 "out|useful|i_therm|solar_i|M1",
                                                 "out|useful_steel|lt_heat|solar_steel|*",
                                                 "out|useful_aluminum|lt_heat|solar_aluminum|*",
                                                 "out|useful_cement|lt_heat|solar_cement|*",
                                                 "out|useful_petro|lt_heat|solar_petro|*",
                                                 "out|useful_resins|lt_heat|solar_resins|*",
                                                "in|secondary|coal|meth_coal|feedstock",
                                                'in|secondary|coal|meth_coal_ccs|feedstock',
                                                "in|secondary|electr|meth_coal|feedstock",
                                                'in|secondary|electr|meth_coal_ccs|feedstock',
                                                'in|secondary|electr|meth_ng_ccs|feedstock',
                                                "in|secondary|electr|meth_ng|feedstock",
                                                'in|secondary|gas|meth_ng_ccs|feedstock',
                                                "in|secondary|gas|meth_ng|feedstock",
                                                'in|primary|biomass|meth_bio|feedstock',
                                                'in|primary|biomass|meth_bio_ccs|feedstock',
                                                'in|secondary|hydrogen|meth_h2|feedstock',
                                                'in|secondary|electr|meth_h2|feedstock',
                                                 ], inplace=True)

            all_flows = df_final_energy.timeseries().reset_index()
            splitted_vars = [v.split("|") for v in all_flows.variable]
            aux1_df = pd.DataFrame(
                splitted_vars,
                columns=["flow_type", "level", "commodity", "technology", "mode"],
            )
            aux2_df = pd.concat(
                [all_flows.reset_index(drop=True), aux1_df.reset_index(drop=True)],
                axis=1,
            )

            # Include only the related industry sector variables

            var_sectors = [
                v
                for v in aux2_df["variable"].values
                if (
                      (v.split('|')[3].endswith("cement"))
                     | (v.split('|')[3].endswith("steel"))
                     | (v.split('|')[3].endswith("aluminum"))
                     | (v.split('|')[3].endswith("petro"))
                     | (v.split('|')[3].endswith("resins"))
                     | (v.split('|')[3].endswith("_i"))
                     | (v.split('|')[3].endswith("_I"))
                     | (('NH3') in v)
                     | (v.split('|')[3].endswith("_fs"))
                     | (v.split('|')[3].startswith("meth"))
                     | (v.split('|')[3].startswith("CH2O"))
                     )
            ]
            aux2_df = aux2_df[aux2_df["variable"].isin(var_sectors)]

            df_final_energy.filter(variable=var_sectors, inplace=True)

            # Aggregate

            if c == 'solar':
                df_final_energy.aggregate(
                    "Final Energy|Industry|Solar", components=var_sectors, append=True,
                )
                df_final_energy.filter(variable=["Final Energy|Industry|Solar"], inplace=True)
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "all":
                df_final_energy.aggregate(
                    "Final Energy|Industry", components=var_sectors, append=True,
                )
                df_final_energy.filter(variable=["Final Energy|Industry"], inplace=True)
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "electr":
                df_final_energy.aggregate(
                    "Final Energy|Industry|Electricity",
                    components=var_sectors,
                    append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Industry|Electricity"], inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "gas":
                # Can not distinguish by type Gases (natural gas, biomass, synthetic fossil, efuel)
                # (coal_gas), from biomass (gas_bio), natural gas (gas_bal): All go into secondary level
                # Can not be distinguished in the final level.
                df_final_energy.aggregate(
                    "Final Energy|Industry|Gases", components=var_sectors, append=True,
                )

                df_final_energy.filter(
                    variable=["Final Energy|Industry|Gases"], inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

            if c == "hydrogen":
                df_final_energy.aggregate(
                    "Final Energy|Industry|Hydrogen",
                    components=var_sectors,
                    append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Industry|Hydrogen"], inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

            if c == "liquids":
                # All liquids
                df_final_energy.aggregate(
                    "Final Energy|Industry|Liquids",
                    components=var_sectors,
                    append=True,
                )
                # Only bios (ethanol)
                filter_vars = [
                    v for v in aux2_df["variable"].values if (("ethanol" in v)
                    & ("methanol" not in v))
                ]
                df_final_energy.aggregate(
                    "Final Energy|Industry|Liquids|Biomass",
                    components=filter_vars,
                    append=True,
                )

                # Oils
                filter_vars = [
                    v
                    for v in aux2_df["variable"].values
                    if (
                        ("naphtha" in v)
                        | ("atm_gasoil" in v)
                        | ("vacuum_gasoil" in v)
                        | ("fueloil" in v)
                        | ("lightoil" in v)
                    )
                ]
                df_final_energy.aggregate(
                    "Final Energy|Industry|Liquids|Oil",
                    components=filter_vars,
                    append=True,
                )

                # Methanol
                filter_vars = [
                    v
                    for v in aux2_df["variable"].values
                    if (
                         ("methanol" in v)
                    )
                ]

                df_final_energy.aggregate(
                    "Final Energy|Industry|Liquids|Coal",
                    components=filter_vars,
                    append=True,
                )

                # Natural Gas Liquids (Ethane/Propane)

                filter_vars = [
                    v
                    for v in aux2_df["variable"].values
                    if (
                        ("gas_proc" in v)
                    )
                ]

                df_final_energy.aggregate(
                    "Final Energy|Industry|Liquids|Gas",
                    components=filter_vars,
                    append=True,
                )

                df_final_energy.filter(
                    variable=[
                        "Final Energy|Industry|Liquids",
                        "Final Energy|Industry|Liquids|Oil",
                        "Final Energy|Industry|Liquids|Biomass",
                        "Final Energy|Industry|Liquids|Coal",
                        "Final Energy|Industry|Liquids|Gas"
                    ],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "solids":
                # All
                df_final_energy.aggregate(
                    "Final Energy|Industry|Solids", components=var_sectors, append=True,
                )

                # Bio
                filter_vars = [
                    v for v in aux2_df["variable"].values if ("biomass" in v)
                ]

                df_final_energy.aggregate(
                    "Final Energy|Industry|Solids|Biomass",
                    components=filter_vars,
                    append=True,
                )

                # Fossil
                filter_vars = [
                    v for v in aux2_df["variable"].values if ("coal" in v)
                ]

                df_final_energy.aggregate(
                    "Final Energy|Industry|Solids|Coal",
                    components=filter_vars,
                    append=True,
                )
                df_final_energy.filter(
                    variable=[
                        "Final Energy|Industry|Solids",
                        "Final Energy|Industry|Solids|Biomass",
                        "Final Energy|Industry|Solids|Coal",
                    ],
                    inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)
            if c == "heat":
                df_final_energy.aggregate(
                    "Final Energy|Industry|Heat", components=var_sectors, append=True,
                )
                df_final_energy.filter(
                    variable=["Final Energy|Industry|Heat"], inplace=True,
                )
                df_final_energy.convert_unit(
                    "GWa", to="EJ/yr", factor=0.03154, inplace=True
                )
                df_final.append(df_final_energy, inplace=True)

    # FINAL ENERGY BY SECTOR AND FUEL
    # Feedstock not included

    sectors = [
        "aluminum",
        "steel",
        "cement",
        "petro",
        "Non-Ferrous Metals",
        "Non-Metallic Minerals",
        "Chemicals",
        'Other Sector'
    ]
    print("Final Energy (excl non-energy use) by sector and fuel is being printed")
    for r in nodes:
        for s in sectors:
            df_final_energy = df.copy()
            df_final_energy.convert_unit("", to="GWa", factor=1, inplace=True)
            exclude = [
                "in|final|atm_gasoil|steam_cracker_petro|*",
                "in|final|ethanol|ethanol_to_ethylene_petro|M1",
                "in|final|gas|gas_processing_petro|M1",
                "in|final|naphtha|steam_cracker_petro|*",
                "in|final|vacuum_gasoil|steam_cracker_petro|*",
                "in|final|*|cokeoven_steel|*",
                "in|final|bf_gas|*",
                "in|final|co_gas|*"]

            df_final_energy.filter(region=r, year=years, inplace=True)
            df_final_energy.filter(variable=["in|final|*",
             'out|useful|i_therm|solar_i|M1',
             'out|useful_steel|lt_heat|solar_steel|low_temp',
             'out|useful_aluminum|lt_heat|solar_aluminum|low_temp',
             'out|useful_cement|lt_heat|solar_cement|low_temp',
             'out|useful_petro|lt_heat|solar_petro|low_temp',
             'out|useful_resins|lt_heat|solar_resins|low_temp',
             'in|secondary|electr|NH3_to_N_fertil|M1',
             'in|secondary|electr|coal_NH3|M1',
             'in|secondary|electr|fueloil_NH3|M1',
             'in|secondary|electr|gas_NH3|M1',
             'in|secondary|electr|coal_NH3_ccs|M1',
             'in|secondary|electr|fueloil_NH3_ccs|M1',
             'in|secondary|electr|gas_NH3_ccs|M1',
             'in|secondary|electr|biomass_NH3_ccs|M1',
             'in|secondary|electr|biomass_NH3|M1',
             'in|secondary|electr|meth_ng|feedstock',
             'in|secondary|electr|meth_ng_ccs|feedstock',
              'in|secondary|electr|meth_coal|feedstock',
              'in|secondary|electr|meth_coal_ccs|feedstock',
             'in|secondary|electr|meth_h2|feedstock'
            ], inplace=True)
            df_final_energy.filter(variable=exclude, keep=False, inplace=True)

            # Decompose the pyam table into pandas data frame

            all_flows = df_final_energy.timeseries().reset_index()

            # Split the strings in the identified variables for further processing
            splitted_vars = [v.split("|") for v in all_flows.variable]

            # Create auxilary dataframes for processing
            aux1_df = pd.DataFrame(
                splitted_vars,
                columns=["flow_type", "level", "commodity", "technology", "mode"],
            )
            aux2_df = pd.concat(
                [all_flows.reset_index(drop=True), aux1_df.reset_index(drop=True)],
                axis=1,
            )

            # To be able to report the higher level sectors.
            if s == "Non-Ferrous Metals":
                tec = [t for t in aux2_df["technology"].values if "aluminum" in t]
                solar_tec = ['solar_aluminum']
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            elif s == "Non-Metallic Minerals":
                tec = [t for t in aux2_df["technology"].values if "cement" in t]
                solar_tec = ['solar_cement']
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            elif s == "Chemicals":
                tec = [t for t in aux2_df["technology"].values if (("petro" in t)
                                                                    | ('NH3' in t)
                                                                    | ( t.startswith('meth_')
                                                                    & (not (t.startswith('meth_i'))))
                                                                    |  ('CH2O'in t)
                                                                    |  ("resins" in t))]
                solar_tec = ['solar_petro', 'solar_resins']
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            elif s == 'Other Sector':
                tec = [t for t in aux2_df["technology"].values if (
                    ((t.endswith("_i"))
                     | (t.endswith('_I'))))]
                solar_tec = ['solar_i']
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            else:
                # Filter the technologies only for the certain industry sector
                tec = [t for t in aux2_df["technology"].values if s in t]
                solar_tec = ['solar_' + s]
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]

            s = change_names(s)

            # Lists to keep commodity, aggregate and variable names.

            commodity_list = []
            aggregate_list = []
            var_list = []

            # For the categoris below filter the required variable names,
            # create a new aggregate name

            commodity_list = [
                "electr",
                "gas",
                "hydrogen",
                "liquids",
                "liquid_bio",
                "liquid_fossil",
                'liquid_other',
                "solids",
                "solids_bio",
                "solids_fossil",
                "heat",
                'solar',
                "all",
            ]

            for c in commodity_list:
                if c == "electr":
                    var = np.unique(
                        aux2_df.loc[aux2_df["commodity"] == "electr", "variable"].values
                    ).tolist()

                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Electricity"
                    )
                elif c == "gas":
                    var = np.unique(
                        aux2_df.loc[aux2_df["commodity"] == "gas", "variable"].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|" + s + "|" + "Gases"
                    )
                elif c == 'solar':
                    var = np.unique(
                        aux2_df.loc[
                            aux2_df["technology"].isin(solar_tec), "variable"
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Solar"
                    )
                elif c == "hydrogen":
                    var = np.unique(
                        aux2_df.loc[
                            aux2_df["commodity"] == "hydrogen", "variable"
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Hydrogen"
                    )
                elif c == "liquids":
                    var = np.unique(
                        aux2_df.loc[
                            (
                                (aux2_df["commodity"] == "fueloil")
                                | (aux2_df["commodity"] == "lightoil")
                                | (aux2_df["commodity"] == "methanol")
                                | (aux2_df["commodity"] == "ethanol")
                            ),
                            "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Liquids"
                    )
                elif c == "liquid_bio":
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "ethanol")), "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Liquids|Biomass"
                    )
                elif c == "liquid_fossil":
                    var = np.unique(
                        aux2_df.loc[
                            (
                                (aux2_df["commodity"] == "fueloil")
                                | (aux2_df["commodity"] == "lightoil")
                            ),
                            "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Liquids|Oil"
                    )
                elif c == "liquid_other":
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "methanol")),
                            "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Liquids|Coal"
                    )
                elif c == "solids":
                    var = np.unique(
                        aux2_df.loc[
                            (
                                (aux2_df["commodity"] == "coal")
                                | (aux2_df["commodity"] == "biomass")
                                | (aux2_df["commodity"] == "coke_iron")
                            ),
                            "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Solids"
                    )
                elif c == "solids_bio":
                    var = np.unique(
                        aux2_df.loc[
                            (aux2_df["commodity"] == "biomass"), "variable"
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Solids|Biomass"
                    )
                elif c == "solids_fossil":
                    var = np.unique(
                        aux2_df.loc[
                            (
                                (aux2_df["commodity"] == "coal")
                                | (aux2_df["commodity"] == "coke_iron")
                            ),
                            "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|"
                        + s
                        + "|"
                        + "Solids|Coal"
                    )
                elif c == "heat":
                    var = np.unique(
                        aux2_df.loc[
                            (aux2_df["commodity"] == "d_heat"), "variable"
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry excl Non-Energy Use|" + s + "|" + "Heat"
                    )
                elif c == "all":
                    var = aux2_df["variable"].tolist()
                    aggregate_name = "Final Energy|Industry excl Non-Energy Use|" + s

                aggregate_list.append(aggregate_name)
                var_list.append(var)

            # Obtain the iamc format dataframe again

            aux2_df.drop(
                ["flow_type", "level", "commodity", "technology", "mode"],
                axis=1,
                inplace=True,
            )
            if aux2_df.size == 0:
                continue
            df_final_energy = pyam.IamDataFrame(data=aux2_df)

            # Aggregate the commodities in iamc object

            for i, c in enumerate(commodity_list):
                if var_list[i]:
                    df_final_energy.aggregate(
                        aggregate_list[i], components=var_list[i], append=True
                    )

            df_final_energy.filter(variable=aggregate_list, inplace=True)
            df_final_energy.convert_unit(
                "GWa", to="EJ/yr", factor=0.03154, inplace=True
            )
            df_final.append(df_final_energy, inplace=True)

    # FINAL ENERGY (NON-ENERGY USE) BY SECTOR AND FUEL
    # Only for high value chemcials there is non-energy use reported.
    # (not in aluminum, steel, cement).

    sectors = ["petro",'ammonia','methanol','Chemicals|Other Sector']
    print("Final Energy non-energy use by sector and fuel is being printed")
    for r in nodes:
        for s in sectors:
            df_final_energy = df.copy()
            df_final_energy.convert_unit("", to="GWa", factor=1, inplace=True)
            include = [
                "in|final|atm_gasoil|steam_cracker_petro|*",
                "in|final|ethanol|ethanol_to_ethylene_petro|M1",
                "in|final|gas|gas_processing_petro|M1",
                "in|final|naphtha|steam_cracker_petro|*",
                "in|final|vacuum_gasoil|steam_cracker_petro|*",
                'in|secondary|electr|electr_NH3|M1',
                'in|secondary|gas|gas_NH3|M1',
                'in|secondary|gas|gas_NH3_ccs|M1',
                'in|secondary|fueloil|fueloil_NH3|M1',
                'in|secondary|fueloil|fueloil_NH3_ccs|M1',
                 "in|secondary|coal|coal_NH3|M1",
                 "in|secondary|coal|coal_NH3_ccs|M1",
                 'in|primary|biomass|biomass_NH3_ccs|M1',
                 'in|primary|biomass|biomass_NH3|M1',
                 "in|final_material|methanol|MTO_petro|energy",
                 "in|final_material|methanol|CH2O_synth|energy",
                 'in|secondary|coal|meth_coal|feedstock',
                 'in|secondary|coal|meth_coal_ccs|feedstock',
                 'in|secondary|gas|meth_ng|feedstock',
                 'in|secondary|gas|meth_ng_ccs|feedstock',
                 'in|primary|biomass|meth_bio|feedstock',
                 'in|primary|biomass|meth_bio_ccs|feedstock',
                 'in|secondary|hydrogen|meth_h2|feedstock',
            ]

            df_final_energy.filter(region=r, year=years, inplace=True)
            df_final_energy.filter(variable=include, inplace=True)

            # Decompose the pyam table into pandas data frame

            all_flows = df_final_energy.timeseries().reset_index()

            # Split the strings in the identified variables for further processing
            splitted_vars = [v.split("|") for v in all_flows.variable]

            # Create auxiliary dataframes for processing
            aux1_df = pd.DataFrame(
                splitted_vars,
                columns=["flow_type", "level", "commodity", "technology", "mode"],
            )
            aux2_df = pd.concat(
                [all_flows.reset_index(drop=True), aux1_df.reset_index(drop=True)],
                axis=1,
            )

            # Filter the technologies only for the certain industry sector
            if s == 'petro':
                tec = [t for t in aux2_df["technology"].values if (s in t)]
            if s == 'ammonia':
                tec = [t for t in aux2_df["technology"].values if ('NH3' in t)]
            if s == 'methanol':
                tec = [t for t in aux2_df["technology"].values if ('meth_' in t)]
            if s == 'Chemicals|Other Sector':
                tec = [t for t in aux2_df["technology"].values if ('CH2O_synth' in t)]


            aux2_df = aux2_df[aux2_df["technology"].isin(tec)]

            s = change_names(s)

            # Lists to keep commodity, aggregate and variable names.

            aggregate_list = []
            commodity_list = []
            var_list = []

            # For the categories below filter the required variable names,
            # create a new aggregate name

            commodity_list = [
                "gas",
                "liquids",
                "liquid_bio",
                'liquid_oil',
                "liquid_gas",
                'methanol',
                "all",
                'solids',
                'solid_coal',
                'solid_bio',
                'electr_gas',
                'hydrogen'
            ]

            for c in commodity_list:
                if c == "electr_gas":
                    var = np.unique(
                        aux2_df.loc[aux2_df["commodity"] == "electr", "variable"].values
                    ).tolist()
                    aggregate_name = "Final Energy|Non-Energy Use|" + s + "|" + "Gases|Electricity"
                elif c == "gas":
                    var = np.unique(
                        aux2_df.loc[(((aux2_df["commodity"] == "gas") | (aux2_df["technology"] == "electr_NH3"))
                        & (aux2_df["technology"] != "gas_processing_petro")) , "variable"].values
                    ).tolist()
                    aggregate_name = "Final Energy|Non-Energy Use|" + s + "|" + "Gases"
                elif c == "hydrogen":
                    var = np.unique(
                        aux2_df.loc[((aux2_df["commodity"] == "hydrogen")
                        & (aux2_df["technology"] == "meth_h2")),"variable"].values
                    ).tolist()
                    aggregate_name = "Final Energy|Non-Energy Use|" + s + "|" + "Hydrogen"
                elif c == "methanol":
                    var = np.unique(
                        aux2_df.loc[
                            (
                            (aux2_df["commodity"] == "methanol")
                            ),
                            "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                            "Final Energy|Non-Energy Use|" + s + "|" + "Other"
                    )
                elif c == "liquids":
                    var = np.unique(
                        aux2_df.loc[
                            (
                                    (aux2_df["commodity"] == "naphtha")
                                    | (aux2_df["commodity"] == "atm_gasoil")
                                    | (aux2_df["commodity"] == "vacuum_gasoil")
                                    | (aux2_df["commodity"] == "ethanol")
                                    | (aux2_df["commodity"] == "fueloil")
                                    | (aux2_df["technology"] == "gas_processing_petro")
                            ),
                            "variable",
                        ].values
                    ).tolist()

                    aggregate_name = (
                            "Final Energy|Non-Energy Use|" + s + "|" + "Liquids"
                    )
                elif c == "liquid_bio":
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "ethanol")), "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                            "Final Energy|Non-Energy Use|" + s + "|" + "Liquids|Biomass"
                    )
                elif c == "liquid_oil":
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "atm_gasoil") |
                             (aux2_df["commodity"] == "naphtha") |
                             (aux2_df["commodity"] == "vacuum_gasoil") |
                             (aux2_df["commodity"] == "fueloil")
                             ), "variable",
                        ].values
                    ).tolist()

                    aggregate_name = (
                            "Final Energy|Non-Energy Use|" + s + "|" + "Liquids|Oil"
                    )
                elif c == "liquid_gas":
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["technology"] == "gas_processing_petro")
                             ), "variable",
                        ].values
                    ).tolist()

                    aggregate_name = (
                            "Final Energy|Non-Energy Use|" + s + "|" + "Liquids|Gas"
                    )
                elif c == 'solids':
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "coal") |
                             (aux2_df["commodity"] == "biomass")), "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                            "Final Energy|Non-Energy Use|" + s + "|" + "Solids"
                    )
                elif c == 'solid_coal':
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "coal") ), "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                            "Final Energy|Non-Energy Use|" + s + "|" + "Solids|Coal"
                    )
                elif c == 'solid_bio':
                    var = np.unique(
                        aux2_df.loc[
                            (
                             (aux2_df["commodity"] == "biomass")), "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                            "Final Energy|Non-Energy Use|" + s + "|" + "Solids|Biomass"
                    )
                elif c == "all":
                    var = aux2_df["variable"].tolist()
                    aggregate_name = "Final Energy|Non-Energy Use|" + s

                aggregate_list.append(aggregate_name)
                var_list.append(var)

            # Obtain the iamc format dataframe again

            aux2_df.drop(
                ["flow_type", "level", "commodity", "technology", "mode"],
                axis=1,
                inplace=True,
            )
            df_final_energy = pyam.IamDataFrame(data=aux2_df)

            # Aggregate the commodities in iamc object

            for i, c in enumerate(commodity_list):
                if var_list[i]:
                    df_final_energy.aggregate(
                        aggregate_list[i], components=var_list[i], append=True
                    )


            df_final_energy.filter(variable=aggregate_list, inplace=True)
            df_final_energy.convert_unit(
                "GWa", to="EJ/yr", factor=0.03154, inplace=True
            )
            df_final.append(df_final_energy, inplace=True)

    # FINAL ENERGY ALL BY SECTOR AND FUEL

    # For ammonia and methanol, there is no seperation for non-energy vs. energy.

    sectors = ['ammonia', 'methanol', 'aluminum', 'steel', 'cement', 'petro',
               "Non-Ferrous Metals", "Non-Metallic Minerals", "Chemicals", 'Chemicals|Other Sector', 'Other Sector']

    print("Final Energy non-energy and energy use by sector and fuel is being printed")
    for r in nodes:
        for s in sectors:
            df_final_energy = df.copy()
            df_final_energy.convert_unit("", to="GWa", factor=1, inplace=True)

            exclude = [
                "in|final|*|cokeoven_steel|*",
                "in|final|bf_gas|*",
                "in|final|co_gas|*",
                'in|final|*|meth_fc_trp|*',
                'in|final|*|meth_ic_trp|*',
                'in|final|*|meth_rc|*']

            include = [
                'in|secondary|coal|coal_NH3|M1',
                'in|secondary|coal|coal_NH3_ccs|M1',
                'in|secondary|electr|coal_NH3|M1',
                'in|secondary|electr|coal_NH3_ccs|M1',
                'in|secondary|fueloil|fueloil_NH3|M1',
                'in|secondary|fueloil|fueloil_NH3_ccs|M1',
                'in|secondary|electr|fueloil_NH3|M1',
                'in|secondary|electr|fueloil_NH3_ccs|M1',
                'in|secondary|gas|gas_NH3|M1',
                'in|secondary|gas|gas_NH3_ccs|M1',
                'in|secondary|electr|gas_NH3|M1',
                'in|secondary|electr|gas_NH3_ccs|M1',
                'in|primary|biomass|biomass_NH3|M1',
                'in|primary|biomass|biomass_NH3_ccs|M1',
                "in|secondary|electr|biomass_NH3|M1",
                "in|secondary|electr|biomass_NH3_ccs|M1",
                "in|secondary|electr|NH3_to_N_fertil|M1",
                'in|secondary|electr|electr_NH3|M1',
                "in|secondary|coal|meth_coal|feedstock",
                'in|secondary|coal|meth_coal_ccs|feedstock',
                "in|secondary|electr|meth_coal|feedstock",
                'in|secondary|electr|meth_coal_ccs|feedstock',
                'in|secondary|gas|meth_ng_ccs|feedstock',
                "in|secondary|gas|meth_ng|feedstock",
                "in|secondary|electr|meth_ng|feedstock",
                'in|secondary|electr|meth_ng_ccs|feedstock',
                'in|primary|biomass|meth_bio|feedstock',
                'in|primary|biomass|meth_bio_ccs|feedstock',
                'in|secondary|hydrogen|meth_h2|feedstock',
                'in|secondary|electr|meth_h2|feedstock',
                "in|final_material|methanol|MTO_petro|energy",
                'in|final|*',
                'out|useful|i_therm|solar_i|M1',
                'out|useful_steel|lt_heat|solar_steel|*',
                'out|useful_cement|lt_heat|solar_cement|*',
                'out|useful_aluminum|lt_heat|solar_aluminum|*',
                'out|useful_petro|lt_heat|solar_steel|*',
                'out|useful_resins|lt_heat|solar_resins|*'
            ]

            df_final_energy.filter(region=r, year=years, inplace=True)
            df_final_energy.filter(variable=include, inplace=True)
            df_final_energy.filter(variable=exclude, keep=False, inplace=True)

            # Decompose the pyam table into pandas data frame

            all_flows = df_final_energy.timeseries().reset_index()

            # Split the strings in the identified variables for further processing
            splitted_vars = [v.split("|") for v in all_flows.variable]

            # Create auxilary dataframes for processing
            aux1_df = pd.DataFrame(
                splitted_vars,
                columns=["flow_type", "level", "commodity", "technology", "mode"],
            )
            aux2_df = pd.concat(
                [all_flows.reset_index(drop=True), aux1_df.reset_index(drop=True)],
                axis=1,
            )

            # Filter the technologies only for the certain industry sector

            if s == "Non-Ferrous Metals":
                tec = [t for t in aux2_df["technology"].values if "aluminum" in t]
                solar_tec = ['solar_aluminum']
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            elif s == "Non-Metallic Minerals":
                tec = [t for t in aux2_df["technology"].values if "cement" in t]
                solar_tec = ['solar_cement']
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            elif s == "Chemicals":
                tec = [t for t in aux2_df["technology"].values if
                ((("petro" in t) & ("MTO_petro" not in t)) \
                | ( t.startswith('meth_') & (not (t.startswith('meth_i')))) | ('NH3' in t) | \
                ('resins' in t) | ('CH2O_synth' in t))]
                solar_tec = ['solar_petro','solar_resins']
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            elif s == "Chemicals|Other Sector":
                tec = [t for t in aux2_df["technology"].values if (('resins' in t) \
                | ('CH2O_synth' in t) | ('CH2O_to_resin' in t) )]
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
                solar_tec = ['solar_resins']
            elif s == 'Other Sector':
                tec = [t for t in aux2_df["technology"].values if (
                    ((t.endswith("_i"))
                     | (t.endswith('_I'))
                     | (t.endswith('_fs'))))]
                solar_tec = ['solar_i']
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            elif s == 'ammonia':
                tec = [t for t in aux2_df["technology"].values if ('NH3' in t)]
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            elif s == 'methanol':
                tec = [t for t in aux2_df["technology"].values if ( t.startswith('meth_') \
                                                   & (not (t.startswith('meth_i'))))]
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
            else:
                # Filter the technologies only for the certain industry sector
                tec = [t for t in aux2_df["technology"].values if s in t]
                aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
                solar_tec = ['solar_' + s]

            s = change_names(s)

            # Lists to keep commodity, aggregate and variable names.

            aggregate_list = []
            commodity_list = []
            var_list = []

            # For the categoris below filter the required variable names,
            # create a new aggregate name

            commodity_list = [
                "electr",
                "gas",
                "hydrogen",
                "liquids",
                "liquid_bio",
                "liquid_fossil",
                "liquid_gas",
                'liquid_other',
                "solids",
                "solids_bio",
                "solids_fossil",
                "heat",
                'solar',
                "all",
            ]

            for c in commodity_list:
                if c == 'electr':
                    var = np.unique(
                        aux2_df.loc[aux2_df["commodity"] == "electr", "variable"].values
                    ).tolist()
                    aggregate_name = "Final Energy|Industry|" + s + "|" + 'Electricity'
                elif c == "gas":
                    var = np.unique(
                        aux2_df.loc[(aux2_df["commodity"] == "gas") & (aux2_df["technology"] != "gas_processing_petro"), "variable"].values
                    ).tolist()
                    aggregate_name = "Final Energy|Industry|" + s + "|" + "Gases"
                elif c == "solar":
                    var = np.unique(
                        aux2_df.loc[aux2_df["technology"].isin(solar_tec), "variable"].values
                    ).tolist()
                    aggregate_name = "Final Energy|Industry|" + s + "|" + "Solar"
                elif c == "hydrogen":
                    var = np.unique(
                        aux2_df.loc[aux2_df["commodity"] == "hydrogen", "variable"].values
                    ).tolist()
                    aggregate_name = "Final Energy|Industry|" + s + "|" + "Hydrogen"
                elif c == "liquids":
                    var = np.unique(
                        aux2_df.loc[
                            (
                                    (aux2_df["commodity"] == "naphtha")
                                    | (aux2_df["commodity"] == "atm_gasoil")
                                    | (aux2_df["commodity"] == "vacuum_gasoil")
                                    | (aux2_df["commodity"] == "ethanol")
                                    | (aux2_df["commodity"] == "fueloil")
                                    | (aux2_df["commodity"] == "lightoil")
                                    | (aux2_df["commodity"] == "methanol")
                                    | (aux2_df["technology"] == "gas_processing_petro")
                            ),
                            "variable",
                        ].values
                    ).tolist()

                    aggregate_name = (
                            "Final Energy|Industry|" + s + "|" + "Liquids"
                    )
                elif c == "liquid_bio":
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "ethanol")), "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                            "Final Energy|Industry|" + s + "|" + "Liquids|Biomass"
                    )
                elif c == "liquid_fossil":
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "atm_gasoil") |
                             (aux2_df["commodity"] == "naphtha") |
                             (aux2_df["commodity"] == "vacuum_gasoil") |
                             (aux2_df["commodity"] == "fueloil") |
                             (aux2_df["commodity"] == "lightoil")
                             ), "variable",
                        ].values
                    ).tolist()

                    aggregate_name = (
                            "Final Energy|Industry|" + s + "|" + "Liquids|Oil"
                    )
                elif c == "liquid_gas":
                    var = np.unique(
                        aux2_df.loc[
                            (aux2_df["technology"] == "gas_processing_petro"), "variable",
                        ].values
                    ).tolist()

                    aggregate_name = (
                            "Final Energy|Industry|" + s + "|" + "Liquids|Gas"
                    )
                elif c == "liquid_other":
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "methanol")),
                            "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry|"+ s+ "|"+ "Liquids|Other"
                    )
                elif c == 'solids':
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "coal") |
                             (aux2_df["commodity"] == "biomass") |
                             (aux2_df["commodity"] == "coke_iron")
                             ), "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                            "Final Energy|Industry|" + s + "|" + "Solids"
                    )
                elif c == 'solids_bio':
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "biomass")), "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                            "Final Energy|Industry|" + s + "|" + "Solids|Biomass"
                    )
                elif c == 'solids_fossil':
                    var = np.unique(
                        aux2_df.loc[
                            ((aux2_df["commodity"] == "coal") |
                            (aux2_df["commodity"] == "coke_iron")
                            ), "variable",
                        ].values
                    ).tolist()
                    aggregate_name = (
                            "Final Energy|Industry|" + s + "|" + "Solids|Coal"
                    )
                elif c == "heat":
                    var = np.unique(
                        aux2_df.loc[
                            (aux2_df["commodity"] == "d_heat"), "variable"
                        ].values
                    ).tolist()
                    aggregate_name = (
                        "Final Energy|Industry|" + s + "|" + "Heat"
                    )
                elif c == "all":
                    var = aux2_df["variable"].tolist()
                    aggregate_name = "Final Energy|Industry|" + s

                aggregate_list.append(aggregate_name)
                var_list.append(var)

            # Obtain the iamc format dataframe again

            aux2_df.drop(
                ["flow_type", "level", "commodity", "technology", "mode"],
                axis=1,
                inplace=True,
            )
            df_final_energy = pyam.IamDataFrame(data=aux2_df)

            # Aggregate the commodities in iamc object

            i = 0
            for c in commodity_list:
                if var_list[i]:
                    df_final_energy.aggregate(
                        aggregate_list[i], components=var_list[i], append=True
                    )

                i = i + 1

            df_final_energy.filter(variable=aggregate_list, inplace=True)
            df_final_energy.convert_unit(
                "GWa", to="EJ/yr", factor=0.03154, inplace=True
            )
            df_final.append(df_final_energy, inplace=True)

    # EMISSIONS
    # If ammonia/methanol is used as feedstock the emissions are accounted under 'CO2_industry',
    # so as 'demand'. If used as fuel, under 'CO2_transformation'.
    # The CCS technologies deduct negative emissions from the overall CO2.

    sectors = [
        "aluminum",
        "steel",
        "petro",
        "cement",
        'ammonia',
        'methanol',
        "all",
        "Chemicals",
        'Chemicals|Other',
        'Other Sector'
    ]

    # CO2_industry and CO2 are reported by the legacy reporting in order to
    # ensure overall consistency for all emissions. In addition remaining industry
    # _i technologies do not have emissions factors multiplied with efficiencies.
    # This causes the emissions from Other Sector to be reported lower.
    emission_type = [
        # 'CO2_industry',
        "CH4",
        # "CO2",
        "NH3",
        "NOx",
        "CF4",
        "N2O",
        "BCA",
        "CO",
        "OCA",
    ]

    print("Emissions are being printed.")
    for typ in ["demand", "process"]:
        for r in nodes:
            for e in emission_type:
                df_emi = df.copy()
                df_emi.filter(region=r, year=years, inplace=True)

                # Exclude M1 but instead include the share that is used for feedstock.
                # meth_import and meth_export have emission coefficients as well
                # but they are excluded to be consistent with other materials emission
                # reporting.

                exclude = ['emis|CO2_industry|meth_coal|M1',
                            'emis|CO2_industry|meth_ng|M1',
                            'emis|CO2_industry|meth_coal_ccs|M1',
                            'emis|CO2_industry|meth_ng_ccs|M1',
                            'emis|CO2|meth_coal_ccs|M1',
                            'emis|CO2|meth_ng_ccs|M1',
                            'emis|CO2|meth_imp|M1',
                            'emis|CO2|meth_exp|M1'
                            ]

                df_emi.filter(variable= exclude, keep=False, inplace=True)

                # Filter the necessary variables

                # CCS technologies have both CO2 (negative) and CO2_transformation
                # coefficent. CO2 is included here because it is the amount captured
                # from the atmosphere and not related to process emissions.

                # "emis|CO2_industry|biomass_NH3_ccs|*","emis|CO2_industry|gas_NH3_ccs|*",
                # "emis|CO2_industry|coal_NH3_ccs|*","emis|CO2_industry|fueloil_NH3_ccs|*",
                # "emis|CO2_industry|biomass_NH3|*","emis|CO2_industry|gas_NH3|*",
                # "emis|CO2_industry|coal_NH3|*","emis|CO2_industry|fueloil_NH3|*",
                # "emis|CO2_industry|electr_NH3|*",'emis|CO2_industry|meth_coal|feedstock',
                # 'emis|CO2_industry|meth_ng|feedstock','emis|CO2_industry|meth_ng_ccs|feedstock',

                if e == 'CO2_industry':
                    emi_filter = ["emis|CO2|biomass_NH3_ccs|*",
                                  "emis|CO2|gas_NH3_ccs|*",
                                  "emis|CO2|coal_NH3_ccs|*",
                                  "emis|CO2|fueloil_NH3_ccs|*",
                                  'emis|CO2|meth_coal_ccs|feedstock',
                                  'emis|CO2|meth_ng_ccs|feedstock',
                                  'emis|CO2_industry|*']

                    df_emi.filter(variable=emi_filter, inplace=True)
                else:
                    emi_filter = ["emis|" + e + "|*"]

                    # Below variables are not reported under Industrial Process
                    # Emissions. Therefore, they are excluded.
                    exclude = ["emis|CO2|biomass_NH3_ccs|*",
                               "emis|CO2|gas_NH3_ccs|*",
                               "emis|CO2|coal_NH3_ccs|*",
                               "emis|CO2|fueloil_NH3_ccs|*",
                               "emis|CO2|meth_coal_ccs|*",
                               "emis|CO2|meth_ng_ccs|*"
                               ]

                    df_emi.filter(variable=exclude, keep=False, inplace=True)
                    df_emi.filter(variable=emi_filter, inplace=True)

                # Perform some specific unit conversions
                if (e == "CO2") | (e == "CO2_industry"):
                    # From MtC to Mt CO2/yr
                    df_emi.convert_unit('', to="Mt CO2/yr", factor=44/12,
                    inplace=True)
                elif (e == "N2O") | (e == "CF4"):
                    unit = "kt " + e + "/yr"
                    df_emi.convert_unit('', to= unit, factor=1,
                    inplace=True)
                else:
                    e = change_names(e)
                    # From kt/yr to Mt/yr
                    unit = "Mt " + e + "/yr"
                    df_emi.convert_unit("", to= unit, factor=0.001, inplace=True)

                all_emissions = df_emi.timeseries().reset_index()

                # Split the strings in the identified variables for further processing
                splitted_vars = [v.split("|") for v in all_emissions.variable]
                # Lists to later keep the variables and names to aggregate
                var_list = []
                aggregate_list = []

                for s in sectors:
                    # Create auxilary dataframes for processing
                    aux1_df = pd.DataFrame(
                        splitted_vars,
                        columns=["emission", "type", "technology", "mode"],
                    )
                    aux2_df = pd.concat(
                        [
                            all_emissions.reset_index(drop=True),
                            aux1_df.reset_index(drop=True),
                        ],
                        axis=1,
                    )
                    # Distinguish the type of emissions and filter the necessary
                    # technologies based on that.

                    if (typ == "process") & (s == "all") & (e != "CO2_industry"):
                        tec = [
                            t
                            for t in aux2_df["technology"].values
                            if (
                                (
                                    (
                                        ("cement" in t)
                                        | ("steel" in t)
                                        | ("aluminum" in t)
                                        | ("petro" in t)
                                    )
                                    & ("furnace" not in t)
                                )
                            )
                        ]
                    if (typ == "process") & (s != "all"):
                        tec = [
                            t
                            for t in aux2_df["technology"].values
                            if ((s in t) & ("furnace" not in t) & ('NH3' not in t) \
                            & (not (t.startswith('meth_'))))
                        ]
                    if (typ == "demand") & (s == "Chemicals"):
                        tec = [
                            t
                            for t in aux2_df["technology"].values
                            if ( (t.startswith('meth_') & (not (t.startswith('meth_i'))))\
                            | ('NH3' in t) | ('MTO' in t) | ('resins' in t) | \
                            (('petro' in t) & ('furnace' in t)))
                        ]
                    if (typ == "demand") & (s == "Chemicals|Other"):
                        tec = [
                            t
                            for t in aux2_df["technology"].values
                            if ('resins' in t)
                            ]
                    if (typ == "demand") & (s == "Other Sector") & (e != "CO2"):
                        tec = [
                            t
                            for t in aux2_df["technology"].values
                            if  (
                                      (t.startswith("biomass_i"))
                                     | (t.startswith("coal_i"))
                                     | (t.startswith("elec_i"))
                                     | (t.startswith("eth_i"))
                                     | (t.startswith("foil_i"))
                                     | (t.startswith("gas_i"))
                                     | (t.startswith("h2_i"))
                                     | (t.startswith("heat_i"))
                                     | (t.startswith("hp_el_i"))
                                     | (t.startswith("hp_gas_i"))
                                     | (t.startswith("loil_i"))
                                     | (t.startswith("meth_i"))
                                     | (t.startswith("sp_coal_I"))
                                     | (t.startswith("sp_el_I"))
                                     | (t.startswith("sp_eth_I"))
                                     | (t.startswith("sp_liq_I"))
                                     | (t.startswith("sp_meth_I"))
                                     | (t.startswith("h2_fc_I"))

                            )
                        ]

                    if (typ == "demand") & (s == "all") & (e != "CO2"):
                        tec = [
                            t
                            for t in aux2_df["technology"].values
                            if (
                                (
                                 (
                                     ("cement" in t)
                                    | ("steel" in t)
                                    | ("aluminum" in t)
                                    | ("petro" in t)

                                    )
                                    & ("furnace" in t)
                               )

                                    | (
                                              (t.startswith("biomass_i"))
                                             | (t.startswith("coal_i"))
                                             | (t.startswith("elec_i"))
                                             | (t.startswith("eth_i"))
                                             | (t.startswith("foil_i"))
                                             | (t.startswith("gas_i"))
                                             | (t.startswith("h2_i"))
                                             | (t.startswith("heat_i"))
                                             | (t.startswith("hp_el_i"))
                                             | (t.startswith("hp_gas_i"))
                                             | (t.startswith("loil_i"))
                                             | (t.startswith("meth_i"))
                                             | (t.startswith("sp_coal_I"))
                                             | (t.startswith("sp_el_I"))
                                             | (t.startswith("sp_eth_I"))
                                             | (t.startswith("sp_liq_I"))
                                             | (t.startswith("sp_meth_I"))
                                             | (t.startswith("h2_fc_I"))
                                            | ("DUMMY_limestone_supply_cement" in t)
                                            | ("DUMMY_limestone_supply_steel" in t)
                                            | ("eaf_steel" in t)
                                            | ("DUMMY_coal_supply" in t)
                                            | ("DUMMY_gas_supply" in t)
                                            | ("NH3" in t)
                                            | t.startswith('meth_')
                                            | ("MTO" in t)
                                            | ("resins" in t)
                                    )
                            )
                        ]

                    if (typ == "demand") & (s != "all") & (s != "Other Sector")\
                    & (s != "Chemicals") & (s != "Chemicals|Other"):

                        if s == "steel":
                            # Furnaces are not used as heat source for iron&steel
                            # Dummy supply technologies help accounting the emissions
                            # from cokeoven_steel, bf_steel, dri_steel, eaf_steel, sinter_steel.

                            tec = [
                                t
                                for t in aux2_df["technology"].values
                                if (
                                    ("DUMMY_coal_supply" in t)
                                    | ("DUMMY_gas_supply" in t)
                                    | ("DUMMY_limestone_supply_steel" in t)
                                )
                            ]

                        elif s == "cement":
                            tec = [
                                t
                                for t in aux2_df["technology"].values
                                if (
                                    ((s in t) & ("furnace" in t))
                                    | ("DUMMY_limestone_supply_cement" in t)
                                )]
                        elif s == "ammonia":
                            tec = [
                                t
                                for t in aux2_df["technology"].values
                                if (
                                    ('NH3' in t))]
                        elif s == "methanol":
                            tec = [
                                t
                                for t in aux2_df["technology"].values
                                if (
                                    t.startswith('meth_') & (not (t.startswith('meth_i'))))]
                        elif s == 'petro':
                            tec = [
                                t
                                for t in aux2_df["technology"].values
                                if ((('petro' in t) & ("furnace" in t)) | ('MTO' in t))
                            ]
                        else:
                            tec = [
                                t
                                for t in aux2_df["technology"].values
                                if ((s in t) & ("furnace" in t))
                            ]
                    # Adjust the sector names
                    s = change_names(s)

                    aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
                    # If there are no emission types for that setor skip
                    if aux2_df.empty:
                        continue

                    # Add elements to lists for aggregation over emission type
                    # for each sector
                    var = aux2_df["variable"].values.tolist()
                    var_list.append(var)

                    # Aggregate names:
                    if s == "all":
                        if (typ == "demand") & (e != "CO2"):
                            if (e != "CO2_industry"):
                                aggregate_name = (
                                    "Emissions|" + e + "|Energy|Demand|Industry"
                                )
                                aggregate_list.append(aggregate_name)
                            else:
                                aggregate_name = (
                                    "Emissions|" + "CO2" + "|Energy|Demand|Industry"
                                )
                                aggregate_list.append(aggregate_name)
                        if ((typ == "process") & (e != "CO2_industry")) :
                            aggregate_name = "Emissions|" + e + "|Industrial Processes"
                            aggregate_list.append(aggregate_name)
                    else:
                        if ((typ == "demand") & (e != "CO2")):
                            if (e != "CO2_industry"):
                                aggregate_name = (
                                    "Emissions|" + e + "|Energy|Demand|Industry|" + s
                                )
                                aggregate_list.append(aggregate_name)
                            else:
                                aggregate_name = (
                                    "Emissions|"
                                    + "CO2"
                                    + "|Energy|Demand|Industry|"
                                    + s
                                )
                                aggregate_list.append(aggregate_name)
                        if ((typ == "process") & (e != "CO2_industry")):
                            aggregate_name = (
                                "Emissions|" + e + "|Industrial Processes|" + s
                            )
                            aggregate_list.append(aggregate_name)

                # To plot:   Obtain the iamc format dataframe again

                aux2_df = pd.concat(
                    [
                        all_emissions.reset_index(drop=True),
                        aux1_df.reset_index(drop=True),
                    ],
                    axis=1,
                )
                aux2_df.drop(
                    ["emission", "type", "technology", "mode"], axis=1, inplace=True
                )
                df_emi = pyam.IamDataFrame(data=aux2_df)

                # Aggregation over emission type for each sector if there are elements to aggregate

                if len(aggregate_list) != 0:
                    for i in range(len(aggregate_list)):
                        df_emi.aggregate(
                            aggregate_list[i], components=var_list[i], append=True
                        )

                    fig, ax1 = plt.subplots(1, 1, figsize=(10, 10))
                    df_emi.filter(variable=aggregate_list, inplace=True)
                    df_emi.plot.stack(ax=ax1)

                    df_final.append(df_emi, inplace=True)
                    ax1.set_title("Emissions_" + r + "_" + e)
                    ax1.set_ylabel("Mt")
                    ax1.legend(bbox_to_anchor=(0.3, 1))

                    plt.close()
                    pp.savefig(fig)

    # PLOTS
    #
    # # HVC Demand: See if this is correct ....
    #
    # for r in nodes:
    #
    #     fig, ax1 = plt.subplots(1, 1, figsize=(10, 10))
    #
    #     if r != "China*":
    #
    #         df_petro = df.copy()
    #         df_petro.filter(region=r, year=years, inplace=True)
    #         df_petro.filter(
    #             variable=[
    #                 "out|final_material|BTX|*",
    #                 "out|final_material|ethylene|*",
    #                 "out|final_material|propylene|*",
    #             ],
    #             inplace=True,
    #         )
    #
    #         # BTX production
    #         BTX_vars = [
    #             "out|final_material|BTX|steam_cracker_petro|atm_gasoil",
    #             "out|final_material|BTX|steam_cracker_petro|naphtha",
    #             "out|final_material|BTX|steam_cracker_petro|vacuum_gasoil",
    #         ]
    #         df_petro.aggregate("BTX production", components=BTX_vars, append=True)
    #
    #         # Propylene production
    #
    #         propylene_vars = [  # "out|final_material|propylene|catalytic_cracking_ref|atm_gasoil",
    #             # "out|final_material|propylene|catalytic_cracking_ref|vacuum_gasoil",
    #             "out|final_material|propylene|steam_cracker_petro|atm_gasoil",
    #             "out|final_material|propylene|steam_cracker_petro|naphtha",
    #             "out|final_material|propylene|steam_cracker_petro|propane",
    #             "out|final_material|propylene|steam_cracker_petro|vacuum_gasoil",
    #         ]
    #
    #         df_petro.aggregate(
    #             "Propylene production", components=propylene_vars, append=True
    #         )
    #
    #         ethylene_vars = [
    #             "out|final_material|ethylene|ethanol_to_ethylene_petro|M1",
    #             "out|final_material|ethylene|steam_cracker_petro|atm_gasoil",
    #             "out|final_material|ethylene|steam_cracker_petro|ethane",
    #             "out|final_material|ethylene|steam_cracker_petro|naphtha",
    #             "out|final_material|ethylene|steam_cracker_petro|propane",
    #             "out|final_material|ethylene|steam_cracker_petro|vacuum_gasoil",
    #         ]
    #
    #         df_petro.aggregate(
    #             "Ethylene production", components=ethylene_vars, append=True
    #         )
    #
    #         if r != "World":
    #
    #             df_petro.filter(
    #                 variable=[
    #                     "BTX production",
    #                     "Propylene production",
    #                     "Ethylene production",
    #                     "out|final_material|BTX|import_petro|*",
    #                     "out|final_material|propylene|import_petro|*",
    #                     "out|final_material|ethylene|import_petro|*",
    #                 ],
    #                 inplace=True,
    #             )
    #         else:
    #             df_petro.filter(
    #                 variable=[
    #                     "BTX production",
    #                     "Propylene production",
    #                     "Ethylene production",
    #                 ],
    #                 inplace=True,
    #             )
    #
    #         df_petro.plot.stack(ax=ax1)
    #         ax1.legend(
    #             [
    #                 "BTX production",
    #                 "Ethylene production",
    #                 "Propylene production",
    #                 "import_BTX",
    #                 "import_ethylene",
    #                 "import_propylene",
    #             ]
    #         )
    #         ax1.set_title("HVC Production_" + r)
    #         ax1.set_xlabel("Years")
    #         ax1.set_ylabel("Mt")
    #
    #         plt.close()
    #         pp.savefig(fig)

    # # Refinery Products - already commented out
    #
    # for r in nodes:
    #
    #     fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(12, 8))
    #     # fig.tight_layout(pad=15.0)
    #
    #     if r != "China*":
    #
    #         # Fuel oil
    #
    #         df_ref_fueloil = df.copy()
    #         df_ref_fueloil.filter(region=r, year=years, inplace=True)
    #
    #         if r == "World":
    #             df_ref_fueloil.filter(
    #                 variable=["out|secondary|fueloil|agg_ref|*"], inplace=True
    #             )
    #
    #         else:
    #             df_ref_fueloil.filter(
    #                 variable=[
    #                     "out|secondary|fueloil|agg_ref|*",
    #                     "out|secondary|fueloil|foil_imp|*",
    #                 ],
    #                 inplace=True,
    #             )
    #
    #         df_ref_fueloil.stack_plot(ax=ax1)
    #
    #         ax1.legend(
    #             [
    #                 "Atm gas oil",
    #                 "Atm resiude",
    #                 "Heavy fuel oil",
    #                 "Petroleum coke",
    #                 "Vacuum residue",
    #                 "import",
    #             ],
    #             bbox_to_anchor=(0.3, 1),
    #         )
    #         ax1.set_title("Fuel oil mix_" + r)
    #         ax1.set_xlabel("Year")
    #         ax1.set_ylabel("GWa")
    #
    #         # Light oil
    #
    #         df_ref_lightoil = df.copy()
    #         df_ref_lightoil.filter(region=r, year=years, inplace=True)
    #
    #         if r == "World":
    #             df_ref_lightoil.filter(
    #                 variable=["out|secondary|lightoil|agg_ref|*"], inplace=True
    #             )
    #
    #         else:
    #             df_ref_lightoil.filter(
    #                 variable=[
    #                     "out|secondary|lightoil|agg_ref|*",
    #                     "out|secondary|lightoil|loil_imp|*",
    #                 ],
    #                 inplace=True,
    #             )
    #
    #         # ,"out|secondary|lightoil|loil_imp|*"
    #         df_ref_lightoil.stack_plot(ax=ax2)
    #         # df_final.append(df_ref_lightoil, inplace = True)
    #         ax2.legend(
    #             [
    #                 "Diesel",
    #                 "Ethane",
    #                 "Gasoline",
    #                 "Kerosene",
    #                 "Light fuel oil",
    #                 "Naphtha",
    #                 "Refinery gas",
    #                 "import",
    #             ],
    #             bbox_to_anchor=(1, 1),
    #         )
    #         ax2.set_title("Light oil mix_" + r)
    #         ax2.set_xlabel("Year")
    #         ax2.set_ylabel("GWa")
    #
    #         plt.close()
    #         pp.savefig(fig)
    #
    # # Oil production World - Already commented out
    #
    # fig, ax1 = plt.subplots(1, 1, figsize=(8, 8))
    #
    # df_all_oil = df.copy()
    # df_all_oil.filter(region="World", year=years, inplace=True)
    # df_all_oil.filter(
    #     variable=[
    #         "out|secondary|fueloil|agg_ref|*",
    #         "out|secondary|lightoil|agg_ref|*",
    #     ],
    #     inplace=True,
    # )
    # df_all_oil.stack_plot(ax=ax1)
    #
    # ax1.legend(
    #     [
    #         "Atmg_gasoil",
    #         "atm_residue",
    #         "heavy_foil",
    #         "pet_coke",
    #         "vaccum_residue",
    #         "diesel",
    #         "ethane",
    #         "gasoline",
    #         "kerosne",
    #         "light_foil",
    #         "naphtha",
    #         "refinery_gas_a",
    #         "refinery_gas_b",
    #     ],
    #     bbox_to_anchor=(1, 1),
    # )
    # ax1.set_title("Oil production" + r)
    # ax1.set_xlabel("Year")
    # ax1.set_ylabel("GWa")
    #
    # plt.close()
    # pp.savefig(fig)

    # Final Energy by all fuels: See if this plot is correct..

    # Select the sectors
    # sectors = ["aluminum", "steel", "cement", "petro"]
    #
    # for r in nodes:
    #
    #     # For each region create a figure
    #
    #     fig, axs = plt.subplots(nrows=2, ncols=2)
    #     fig.set_size_inches(20, 20)
    #     fig.subplots_adjust(wspace=0.2)
    #     fig.subplots_adjust(hspace=0.5)
    #     fig.tight_layout(pad=20.0)
    #
    #     if r != "China*":
    #
    #         # Specify the position of each sector in the graph
    #
    #         cnt = 1
    #
    #         for s in sectors:
    #
    #             if cnt == 1:
    #                 x_cor = 0
    #                 y_cor = 0
    #
    #             if cnt == 2:
    #                 x_cor = 0
    #                 y_cor = 1
    #
    #             if cnt == 3:
    #                 x_cor = 1
    #                 y_cor = 0
    #
    #             if cnt == 4:
    #                 x_cor = 1
    #                 y_cor = 1
    #
    #             df_final_energy = df.copy()
    #             df_final_energy.convert_unit("", to="GWa", factor=1, inplace=True)
    #             df_final_energy.filter(region=r, year=years, inplace=True)
    #             df_final_energy.filter(variable=["in|final|*"], inplace=True)
    #             df_final_energy.filter(
    #                 variable=["in|final|*|cokeoven_steel|*"], keep=False, inplace=True
    #             )
    #
    #             if s == "petro":
    #                 # Exclude the feedstock ethanol and natural gas
    #                 df_final_energy.filter(
    #                     variable=[
    #                         "in|final|ethanol|ethanol_to_ethylene_petro|M1",
    #                         "in|final|gas|gas_processing_petro|M1",
    #                         "in|final|atm_gasoil|steam_cracker_petro|atm_gasoil",
    #                         "in|final|vacuum_gasoil|steam_cracker_petro|vacuum_gasoil",
    #                         "in|final|naphtha|steam_cracker_petro|naphtha",
    #                     ],
    #                     keep=False,
    #                     inplace=True,
    #                 )
    #
    #             all_flows = df_final_energy.timeseries().reset_index()
    #
    #             # Split the strings in the identified variables for further processing
    #             splitted_vars = [v.split("|") for v in all_flows.variable]
    #
    #             # Create auxilary dataframes for processing
    #             aux1_df = pd.DataFrame(
    #                 splitted_vars,
    #                 columns=["flow_type", "level", "commodity", "technology", "mode"],
    #             )
    #             aux2_df = pd.concat(
    #                 [all_flows.reset_index(drop=True), aux1_df.reset_index(drop=True)],
    #                 axis=1,
    #             )
    #
    #             # Filter the technologies only for the certain material
    #             tec = [t for t in aux2_df["technology"].values if s in t]
    #             aux2_df = aux2_df[aux2_df["technology"].isin(tec)]
    #
    #             # Lists to keep commodity, aggregate and variable.
    #
    #             aggregate_list = []
    #             commodity_list = []
    #             var_list = []
    #
    #             # For each commodity collect the variable name, create an aggregate name
    #             s = change_names(s)
    #             for c in np.unique(aux2_df["commodity"].values):
    #                 var = np.unique(
    #                     aux2_df.loc[aux2_df["commodity"] == c, "variable"].values
    #                 ).tolist()
    #                 aggregate_name = "Final Energy|" + s + "|" + c
    #
    #                 aggregate_list.append(aggregate_name)
    #                 commodity_list.append(c)
    #                 var_list.append(var)
    #
    #             # Obtain the iamc format dataframe again
    #
    #             aux2_df.drop(
    #                 ["flow_type", "level", "commodity", "technology", "mode"],
    #                 axis=1,
    #                 inplace=True,
    #             )
    #             df_final_energy = pyam.IamDataFrame(data=aux2_df)
    #
    #             # Aggregate the commodities in iamc object
    #
    #             i = 0
    #             for c in commodity_list:
    #                 df_final_energy.aggregate(
    #                     aggregate_list[i], components=var_list[i], append=True
    #                 )
    #                 i = i + 1
    #
    #             df_final_energy.convert_unit(
    #                 "GWa", to="EJ/yr", factor=0.03154, inplace=True
    #             )
    #             df_final_energy.filter(variable=aggregate_list).plot.stack(
    #                 ax=axs[x_cor, y_cor]
    #             )
    #             axs[x_cor, y_cor].set_ylabel("EJ/yr")
    #             axs[x_cor, y_cor].set_title("Final Energy_" + s + "_" + r)
    #             cnt = cnt + 1
    #
    #     plt.close()
    #     pp.savefig(fig)

    # Scrap Release: (Buildings), Other and Power Sector
    # For cement, we dont have any other scrap represented in the model.
    # Only scrap is from power sector.
    print('Scrap generated by sector')
    materials = ["aluminum","steel","cement"]

    for r in nodes:
        for m in materials:
            df_scrap_by_sector = df.copy()
            df_scrap_by_sector.filter(region=r, year=years, inplace=True)

            # filt_buildings = 'out|end_of_life|' + m + '|demolition_build|M1'
            # print(filt_buildings)
            filt_other = 'out|end_of_life|' + m + '|other_EOL_' + m + '|M1'
            filt_total = 'in|end_of_life|' + m + '|total_EOL_' + m + '|M1'

            m = change_names(m)
            # var_name_buildings = 'Total Scrap|Residential and Commercial|' + m
            var_name_other = 'Total Scrap|Other|' + m
            var_name_power = 'Total Scrap|Power Sector|' + m

            if m != "Non-Metallic Minerals|Cement":
                df_scrap_by_sector.aggregate(var_name_other,\
                components=[filt_other],append=True)

            # df_scrap_by_sector.aggregate(var_name_buildings,\
            # components=[filt_buildings],append=True)
            # 'out|end_of_life|' + m + '|demolition_build|M1',

                df_scrap_by_sector.subtract(filt_total, filt_other,var_name_power,
                axis='variable', append = True)

                df_scrap_by_sector.filter(variable=[
                # var_name_buildings,
                var_name_other, var_name_power],inplace=True)
            elif m == "Non-Metallic Minerals|Cement":
                df_scrap_by_sector.aggregate(var_name_power,\
                components=[filt_total],append=True)

                df_scrap_by_sector.filter(variable=[
                # var_name_buildings,
                var_name_power],inplace=True)

            df_scrap_by_sector.convert_unit('', to='Mt/yr', factor=1, inplace = True)
            df_final.append(df_scrap_by_sector, inplace=True)

    # PRICE
    # SCRAP_RECOVERY_ALUMINUM NEEDS TO BE FIXED

    # df_final = df_final.timeseries().reset_index()
    #
    # commodity_type = [
    #     "Non-Ferrous Metals|Aluminium",
    #     "Non-Ferrous Metals|Aluminium|New Scrap",
    #     "Non-Ferrous Metals|Aluminium|Old Scrap",
    #     "Non-Ferrous Metals|Bauxite",
    #     "Non-Ferrous Metals|Alumina",
    #     "Steel|Iron Ore",
    #     "Steel|Pig Iron",
    #     "Steel|New Scrap",
    #     "Steel|Old Scrap",
    #     "Steel",
    #     "Non-Metallic Minerals|Cement",
    #     "Non-Metallic Minerals|Limestone",
    #     "Non-Metallic Minerals|Clinker Cement",
    #     "Chemicals|High Value Chemicals",
    # ]
    #
    # for c in commodity_type:
    #     prices = scenario.var("PRICE_COMMODITY")
    #     # Used for calculation of average prices for scraps
    #     output = scenario.par(
    #         "output",
    #         filters={"technology": ["scrap_recovery_aluminum", "scrap_recovery_steel"]},
    #     )
    #     # Differs per sector what to report so more flexible with conditions.
    #     # Store the relevant variables in prices_all
    #
    #     # ALUMINUM
    #     if c == "Non-Ferrous Metals|Bauxite":
    #         continue
    #     if c == "Non-Ferrous Metals|Alumina":
    #         prices_all = prices[
    #             (prices["level"] == "secondary_material")
    #             & (prices["commodity"] == "aluminum")
    #         ]
    #     if c == "Non-Ferrous Metals|Aluminium":
    #         prices_all = prices[
    #             (prices["level"] == "final_material")
    #             & (prices["commodity"] == "aluminum")
    #         ]
    #     if c == "Non-Ferrous Metals|Aluminium|New Scrap":
    #         prices_all = prices[
    #             (prices["level"] == "new_scrap") & (prices["commodity"] == "aluminum")
    #         ]
    #
    #     # IRON AND STEEL
    #     if c == "Steel|Iron Ore":
    #         prices_all = prices[(prices["commodity"] == "ore_iron")]
    #     if c == "Steel|Pig Iron":
    #         prices_all = prices[(prices["commodity"] == "pig_iron")]
    #     if c == "Steel":
    #         prices_all = prices[
    #             (prices["commodity"] == "steel") & (prices["level"] == "final_material")
    #         ]
    #     if c == "Steel|New Scrap":
    #         prices_all = prices[
    #             (prices["commodity"] == "steel") & (prices["level"] == "new_scrap")
    #         ]
    #     # OLD SCRAP (For aluminum and steel)
    #
    #     if (c == "Steel|Old Scrap") | (c == "Non-Ferrous Metals|Aluminium|Old Scrap"):
    #
    #         prices = prices[
    #             (
    #                 (prices["level"] == "old_scrap_1")
    #                 | (prices["level"] == "old_scrap_2")
    #                 | (prices["level"] == "old_scrap_3")
    #             )
    #         ]
    #
    #         if c == "Non-Ferrous Metals|Aluminium|Old Scrap":
    #             output = output[output["technology"] == "scrap_recovery_aluminum"]
    #             prices = prices[(prices["commodity"] == "aluminum")]
    #         if c == "Steel|Old Scrap":
    #             output = output[output["technology"] == "scrap_recovery_steel"]
    #             prices = prices[(prices["commodity"] == "steel")]
    #
    #         prices.loc[prices["level"] == "old_scrap_1", "weight"] = output.loc[
    #             output["level"] == "old_scrap_1", "value"
    #         ].values[0]
    #
    #         prices.loc[prices["level"] == "old_scrap_2", "weight"] = output.loc[
    #             output["level"] == "old_scrap_2", "value"
    #         ].values[0]
    #
    #         prices.loc[prices["level"] == "old_scrap_3", "weight"] = output.loc[
    #             output["level"] == "old_scrap_3", "value"
    #         ].values[0]
    #
    #         prices_all = pd.DataFrame(columns=["node", "commodity", "year"])
    #         prices_new = pd.DataFrame(columns=["node", "commodity", "year"])
    #
    #         for reg in output["node_loc"].unique():
    #             #for yr in output["year_act"].unique():
    #             for yr in prices["year"].unique():
    #                 prices_temp = prices.groupby(["node", "year"]).get_group((reg, yr))
    #                 rate = prices_temp["weight"].values.tolist()
    #                 amount = prices_temp["lvl"].values.tolist()
    #                 weighted_avg = np.average(amount, weights=rate)
    #                 prices_new = pd.DataFrame(
    #                     {"node": reg, "year": yr, "commodity": c, "lvl": weighted_avg,},
    #                     index=[0],
    #                 )
    #                 prices_all = pd.concat([prices_all, prices_new])
    #     # CEMENT
    #
    #     if c == "Non-Metallic Minerals|Limestone":
    #         prices_all = prices[(prices["commodity"] == "limestone_cement")]
    #     if c == "Non-Metallic Minerals|Clinker Cement":
    #         prices_all = prices[(prices["commodity"] == "clinker_cement")]
    #     if c == "Non-Metallic Minerals|Cement":
    #         prices_all = prices[
    #             (prices["commodity"] == "cement") & (prices["level"] == "demand")
    #         ]
    #     # Petro-Chemicals
    #
    #     if c == "Chemicals|High Value Chemicals":
    #         prices = prices[
    #             (prices["commodity"] == "ethylene")
    #             | (prices["commodity"] == "propylene")
    #             | (prices["commodity"] == "BTX")
    #         ]
    #         prices_all = prices.groupby(by=["year", "node"]).mean().reset_index()
    #
    #     # Convert all to IAMC format.
    #     for r in prices_all["node"].unique():
    #         if (r == "R11_GLB") | (r == "R12_GLB"):
    #             continue
    #         df_price = pd.DataFrame(
    #             {"model": model_name, "scenario": scenario_name, "unit": "2010USD/Mt",},
    #             index=[0],
    #         )
    #
    #         for y in prices_all["year"].unique():
    #             df_price["region"] = r
    #             df_price["variable"] = "Price|" + c
    #             x = prices_all.loc[
    #                 ((prices_all["node"] == r) & (prices_all["year"] == y)), "lvl"
    #             ]
    #             if not x.empty:
    #                 value = x.values[0] * 1.10774
    #             else:
    #                 value = 0
    #             df_price[y] = value
    #
    #             df.price = df_price.columns.astype(str)
    #
    #         df_final = pd.concat([df_final, df_price])

    # Material Demand - comment out if no power sector
    # Power Sector

    # input_cap_new = scenario.par("input_cap_new")
    # input_cap_new.drop(
    #     ["node_origin", "level", "time_origin", "unit"], axis=1, inplace=True
    # )
    # input_cap_new.rename(
    #     columns={
    #         "value": "material_intensity",
    #         "node_loc": "region",
    #         "year_vtg": "year",
    #     },
    #     inplace=True,
    # )
    #
    # cap_new = scenario.var("CAP_NEW")
    # cap_new.drop(["mrg"], axis=1, inplace=True)
    # cap_new.rename(
    #     columns={"lvl": "installed_capacity", "node_loc": "region", "year_vtg": "year"},
    #     inplace=True,
    # )
    #
    # merged_df = pd.merge(cap_new, input_cap_new)
    # merged_df["Material Need"] = (
    #     merged_df["installed_capacity"] * merged_df["material_intensity"]
    # )
    # merged_df = merged_df[merged_df["year"] >= min(years)]
    #
    # final_material_needs = (
    #     merged_df.groupby(["commodity", "region", "year"])
    #     .sum()
    #     .drop(["installed_capacity", "material_intensity"], axis=1)
    # )
    #
    # final_material_needs = final_material_needs.reset_index(["year"])
    # final_material_needs = pd.pivot_table(
    #     final_material_needs,
    #     values="Material Need",
    #     columns="year",
    #     index=["commodity", "region"],
    # ).reset_index(["commodity", "region"])
    #
    # final_material_needs_global = (
    #     final_material_needs.groupby("commodity").sum().reset_index()
    # )
    # final_material_needs_global["region"] = "World"
    #
    # material_needs_all = pd.concat(
    #     [final_material_needs, final_material_needs_global], ignore_index=True
    # )
    #
    # material_needs_all["scenario"] = scenario_name
    # material_needs_all["model"] = model_name
    # material_needs_all["unit"] = "Mt/yr"
    # material_needs_all["commodity"] = material_needs_all.apply(
    #     lambda x: change_names(x["commodity"]), axis=1
    # )
    # material_needs_all = material_needs_all.assign(
    #     variable=lambda x: "Material Demand|Power Sector|" + x["commodity"]
    # )
    #
    # material_needs_all = material_needs_all.drop(["commodity"], axis=1)
    # df.price = df_price.columns.astype(str)
    #
    # print("This is the final_material_needs")
    # print(material_needs_all)
    #
    # df_final = pd.concat([df_final, material_needs_all])

    # Trade
    # ....................

    path_temp = os.path.join(directory, "temp_new_reporting.xlsx")
    df_final.to_excel(path_temp, sheet_name="data", index=False)

    excel_name_new = "New_Reporting_" + model_name + "_" + scenario_name + ".xlsx"
    path_new = os.path.join(directory, excel_name_new)

    print(path_temp)
    print(path_new)
    fix_excel(path_temp, path_new)
    print("New reporting file generated.")
    df_final = pd.read_excel(path_new)

    # df_resid = pd.read_csv(path_resid)
    # df_resid["Model"] = model_name
    # df_resid["Scenario"] = scenario_name
    # df_comm = pd.read_csv(path_comm)
    # df_comm["Model"] = model_name
    # df_comm["Scenario"] = scenario_name

    scenario.check_out(timeseries_only=True)
    print("Starting to upload timeseries")
    print(df_final.head())
    scenario.add_timeseries(df_final)
    # scenario.add_timeseries(df_resid)
    # scenario.add_timeseries(df_comm)
    print("Finished uploading timeseries")
    scenario.commit("Reporting uploaded as timeseries")

    pp.close()
    os.remove(path_temp)
    #os.remove(path)
