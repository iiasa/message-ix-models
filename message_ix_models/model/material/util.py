import os
from pathlib import Path
from typing import Any, Union

import message_ix
import openpyxl as pxl
import pandas as pd
import pycountry
import yaml
from scipy.optimize import curve_fit

from message_ix_models import Context
from message_ix_models.util import load_package_data, package_data_path

# Configuration files
METADATA = [
    # ("material", "config"),
    ("material", "set"),
    # ("material", "technology"),
]


def read_config() -> Context:
    """Read configuration from set.yaml.

    Returns
    -------
    message_ix_models.Context
        Context object holding information about MESSAGEix-Materials structure
    """
    # TODO this is similar to transport.utils.read_config; make a common
    #      function so it doesn't need to be in this file.
    context = Context.get_instance(-1)

    if "material set" in context:
        # Already loaded
        return context

    # Load material configuration
    for parts in METADATA:
        # Key for storing in the context
        key = " ".join(parts)

        # Actual filename parts; ends with YAML
        _parts = list(parts)
        _parts[-1] += ".yaml"

        context[key] = load_package_data(*_parts)

    # Use a shorter name
    context["material"] = context["material set"]
    return context


def prepare_xlsx_for_explorer(filepath: str) -> None:
    """
    Post-processing helper to make reporting files compliant for
    upload to IIASA Scenario Explorer

    Parameters
    ----------
    filepath : str
        Path to xlsx files generated with message_ix_models.report.legacy

    """
    df = pd.read_excel(filepath)

    def add_R12(str: str) -> str:
        return "R12_" + str if len(str) < 5 else str

    df = df[~df["Region"].isna()]
    df["Region"] = df["Region"].map(add_R12)
    df.to_excel(filepath, index=False)


def combine_df_dictionaries(*args: dict[str, pd.DataFrame]) -> dict:
    """
    Iterates through dictionary items and collects all values with same keys
     from dictionaries in one dict
    Parameters
    ----------
    args: dict[str, pd.DataFrame]
        arbitrary number of dictionaries with str keys and pd.DataFrame values

    Returns
    -------
    dict[str, pd.DataFrame]
        dictionary containing all unique elements of
        pd.DataFrames provided by *args dict
    """
    keys = set([key for tup in args for key in tup])
    comb_dict = {}
    for i in keys:
        comb_dict[i] = pd.concat([j.get(i) for j in args])
    return comb_dict


def read_yaml_file(file_path: Union[str, Path]) -> Union[dict, None]:
    """
    Tries to read yaml file into a dict

    Parameters
    ----------
    file_path : str
        file path to yaml file

    Returns
    -------
    dict
    """
    with open(file_path, encoding="utf8") as file:
        try:
            data = yaml.safe_load(file)
            return data
        except yaml.YAMLError as e:
            print(f"Error while parsing YAML file: {e}")
            return None


# NOTE guessing the type hint here, but this seems unused anyway
def invert_dictionary(original_dict: dict[str, list[str]]) -> dict[str, list[str]]:
    """
    Create inverted dictionary from existing dictionary, where values turn
    into keys and vice versa

    Parameters
    ----------
    original_dict: dict
        dictionary with values of list type

    Returns
    -------
    dict

    """
    inverted_dict: dict[str, list[str]] = {}
    for key, value in original_dict.items():
        for array_element in value:
            if array_element not in inverted_dict:
                inverted_dict[array_element] = []
            inverted_dict[array_element].append(key)
    return inverted_dict


def excel_to_csv(material_dir: str, fname: str) -> None:
    """
    Helper to create trackable copies xlsx files used for MESSAGEix-Materials
    data input by printing each sheet to a csv file. Output is saved in
     "data/materials/version control"

    Parameters
    ----------
    material_dir : str
        path to industry sector data folder
    fname : str
        file name of xlsx file
    """
    xlsx_dict = pd.read_excel(
        package_data_path("material", material_dir, fname), sheet_name=None
    )
    if not os.path.isdir(package_data_path("material", "version control")):
        os.mkdir(package_data_path("material", "version control"))
    os.mkdir(package_data_path("material", "version control", fname))
    for tab in xlsx_dict.keys():
        xlsx_dict[tab].to_csv(
            package_data_path("material", "version control", fname, f"{tab}.csv"),
            index=False,
        )


def get_all_input_data_dirs() -> list[str]:
    """
    Iteratable for getting all material input data folders

    Returns
    -------
    list of folder names of material data
    """
    elements = os.listdir(package_data_path("material"))
    elements = [i for i in elements if os.path.isdir(package_data_path("material", i))]
    return elements


def remove_from_list_if_exists(element: Any, _list: list) -> None:
    """
    Utility function removing element from list if it is part of the list
    Parameters
    ----------
    element
        element to remove
    _list
        list that poetntially contains element
    """
    if element in _list:
        _list.remove(element)


def exponential(x: Union[float, list[float]], b: float, m: float) -> float:
    """
    Mathematical function used in Excels GROWTH function

    Parameters
    ----------
    x: float or list
        domain of function
    b : float
        function parameter b
    m: float
        function parameter m
    Returns
    -------
    float
        function value for given b, m and x
    """
    return b * m**x


def price_fit(df: pd.DataFrame) -> float:
    """
    Python implementation of price_ref parameter estimation implemented in
     MESSAGEix-MACRO calibration files.

    Parameters
    ----------
    df: pd.DataFrame
        DataFrame with required columns: "year" and "lvl"
    Returns
    -------
    float
        estimated value for price_ref in 2020
    """

    pars = curve_fit(exponential, df.year, df.lvl, maxfev=10000)[0]
    val = exponential([2020], *pars)[0]
    # print(df.commodity.unique(), df.node.unique(), val)
    return val


def cost_fit(df: pd.DataFrame) -> float:
    """
    Python implementation of cost_ref parameter estimation implemented in
     MESSAGEix-MACRO calibration files.

    Parameters
    ----------
    df: pd.DataFrame
        DataFrame with required columns: "year" and "lvl"
    Returns
    -------
    float
        estimated value for cost_ref in 2020
    """
    # print(df.lvl)
    try:
        pars = curve_fit(exponential, df.year, df.lvl, maxfev=5000)[0]
        val = exponential([2020], *pars)[0]
    except RuntimeError:
        val = df.lvl.values[0]
    # print(df.node.unique(), val)
    return val / 1000


def update_macro_calib_file(scenario: message_ix.Scenario, fname: str) -> None:
    """Function to automate manual steps in MACRO calibration

    Tries to open a xlsx file with the given "fname" and
    writes cost_ref and price_ref values derived from scenario
    "COST_NODAL_NET" and PRICE_COMMODITY" variables to the respective xlsx sheets.

    Parameters
    ----------
    scenario: message_ix.Scenario
        Scenario instance to be calibrated
    fname : str
        file name of MACRO file used for calibration
    """
    # Change this according to the relevant data path
    path = package_data_path("material", "macro", fname)
    wb = pxl.load_workbook(path)

    fmy = scenario.firstmodelyear
    nodes = [
        "R12_AFR",
        "R12_CHN",
        "R12_EEU",
        "R12_FSU",
        "R12_LAM",
        "R12_MEA",
        "R12_NAM",
        "R12_PAO",
        "R12_PAS",
        "R12_RCPA",
        "R12_SAS",
        "R12_WEU",
    ]

    # cost_ref
    years_cost = [i for i in range(fmy, fmy + 15, 5)]
    df = scenario.var("COST_NODAL_NET", filters={"year": years_cost})
    df["node"] = pd.Categorical(df["node"], nodes)
    df = df[df["year"].isin(years_cost)].groupby(["node"]).apply(cost_fit)
    ws = wb["cost_ref"]
    # write derived values to sheet. Cell B7 (MEA region) is skipped.
    for i in range(2, 7):
        ws[f"B{i}"].value = df.values[i - 2]
    for i in range(8, 14):
        ws[f"B{i}"].value = df.values[i - 2]

    # price_ref
    comms = ["i_feed", "i_spec", "i_therm", "rc_spec", "rc_therm", "transport"]
    years_price = [i for i in range(fmy, 2055, 5)]
    df = scenario.var(
        "PRICE_COMMODITY", filters={"commodity": comms, "year": years_price}
    )
    df["node"] = pd.Categorical(df["node"], nodes)
    df["commodity"] = pd.Categorical(df["commodity"], comms)
    df = df.groupby(["node", "commodity"]).apply(price_fit)
    ws = wb["price_ref"]
    for i in range(2, 62):
        ws[f"C{i}"].value = df.values[i - 2]
    wb.save(path)


def get_ssp_from_context(context: Context) -> str:
    """Get selected SSP from context

    Parameters
    ----------
    context: Context

    Returns
    -------
    str
        SSP label
    """
    return "SSP2" if "ssp" not in context else context["ssp"]


def maybe_remove_water_tec(scenario: message_ix.Scenario, results: dict) -> None:
    if len(scenario.par("output", filters={"technology": "extract_surfacewater"})):
        results["input"] = results["input"].replace({"freshwater_supply": "freshwater"})


def path_fallback(context_or_regions: Union[Context, str], *parts) -> Path:
    """Return a :class:`.Path` constructed from `parts`.

    If ``context.model.regions`` (or a string value as the first argument) is defined
    and the file exists in a subdirectory of :file:`data/transport/{regions}/`, return
    its path; otherwise, return the path in :file:`data/transport/`.
    """
    if isinstance(context_or_regions, str):
        regions = context_or_regions
    else:
        # Use a value from a Context object, or a default
        regions = context_or_regions.model.regions

    candidates = (
        package_data_path("material", regions, *parts),
        package_data_path("material", *parts),
    )

    for c in candidates:
        if c.exists():
            return c

    raise FileNotFoundError(candidates)


def get_pycountry_iso(row, mis_dict):
    try:
        row = pycountry.countries.lookup(row).alpha_3
    except LookupError:
        try:
            row = mis_dict[row]
        except KeyError:
            print(f"{row} is not mapped to an ISO")
            row = None
    return row


def get_r12_reg(df, r12_map_inv, col_name):
    try:
        df = r12_map_inv[df[col_name]]
    except KeyError:
        df = None
    return df


def add_R12_column(df, file_path, iso_column="COUNTRY"):
    # Replace 'your_file_path.yaml' with the path to your actual YAML file
    # file_path = private_data_path("node", "R12_SSP_V1.yaml")
    yaml_data = read_yaml_file(file_path)
    yaml_data.pop("World")

    r12_map = {k: v["child"] for k, v in yaml_data.items()}
    r12_map_inv = {k: v[0] for k, v in invert_dictionary(r12_map).items()}

    df["R12"] = df.apply(lambda x: get_r12_reg(x, r12_map_inv, iso_column), axis=1)
    return df
